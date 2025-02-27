import os
import re
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from email.mime.image import MIMEImage
import smtplib

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# Função para remover "CP" e os números após "CP"
def remover_cp(valor):
    if isinstance(valor, str):  # Verificar se o valor é uma string
        # Usar expressão regular para remover "CP" seguido de números
        return re.sub(r"CP\d+", "", valor).strip()  # Remove "CP" e os números
    return valor  # Se não for uma string, retorna o valor original

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()

    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select c.nome_cliente, (s.descricao)serviço, m.os, o.placa, orc.honorarios, orc.taxa_detran, (orc.vistoria)validacao, orc.vlr_correio, orc.ipva_total, orc.prim_parcela, orc.seg_parcela, orc.terc_parcela, orc.total_dpvat, orc.desconto, orc.placas, orc.lacre, orc.outros, orc.srpr, orc.lacracao,
orc.placas_esp, orc.licenciamento, orc.multas, orc.vlr_bo, orc.total, orc.txvistoriamovel, (cp.ipva)ipva_cp, (cp.licenciamento)lic_cp, (cp.multas)multas_cp, codigo_cp, (orc.extras)Reembolso, m.observacao, m.solicitante
from os o
left join cliente c on c.codigo_cliente = o.cod_cliente
left join orcamento orc on orc.os = numero_os
left join cp_os on orc.os = num_os
left join contas_pagar cp on cod_cp = codigo_cp
left join manutencao m on m.os = numero_os
left join pendencias_os pdo on m.os = pdo.os
left join pendencias pd on pdo.codigo = pd.codigo
left join servicos s on o.cod_servico = s.codigo_servico

where pdo.codigo=120
and pdo.data_baixa='30-12-1899' """

    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp)
    # Resultado da consulta
    results = cursor.fetchall()

# Em caso de erro obtem e exibe o erro
except Exception as e:
    print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
# Define o caminho completo da planilha
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
# Define o cabeçalho da planilha
dados = {
         'Nome cliente': [],
         'Serviço': [],
         'OS': [],
         'Placa': [],
         'Honorarios': [],
         'Taxa Detran': [],
         'Validação': [],
         'Valor correio': [],
         'IPVA total': [],
         'Primeira parcela': [],
         'Segunda parcela': [],
         'Terceira parcela': [],
         'Total Dpvat': [],
         'Desconto': [],
         'Placas': [],
         'Lacre': [],
         'Outros': [],
         'Srpr': [],
         'Lacração': [],
         'Placas especial': [],
         'Licenciamento': [],
         'Multas': [],
         'Valor BO': [],
         'Total': [],
         'Taxa vostoria movel': [],
         'IPVA CP': [],
         'Licenciamento CP': [],
         'Multas CP': [],
         'Código CP': [],
         'Reembolso': [],
#         'Descrição': [],
         'Obrservação': [],
         'Solicitante': []
        }

# Crie um DataFrame com os dados
df = pd.DataFrame(dados)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)

wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)

# Atribuir os valores ao cabeçalho
for col, valor in enumerate(dados, start=1):
    ws.cell(row=1, column=col, value=valor)

# Adicionar cor de fundo e estilo ao cabeçalho (linha 1)
cor_fundo = PatternFill(start_color="C3ECBC", end_color="C3ECBC", fill_type="solid")  # Amarelo
font_estilo = Font(bold=True, color="000000")  # Texto em negrito e preto

# Aplicar as mudanças nas células do cabeçalho (linha 1)
for col in range(1, len(dados) + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = cor_fundo  # Cor de fundo
    celula.font = font_estilo  # Estilo da fonte (negrito e preto)
# Salva a planilha após as mudanças    
wb.save(caminho_completo)
# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(caminho_completo)
    ws = wb.active
    # Percorres todas as linhas e acha a linha vazia
    for i, row in enumerate(results, start=primeira_linha_vazia):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
wb.save(caminho_completo)

# Iterar sobre as colunas da planilha (começando da última coluna para evitar problemas ao excluir)
for col in range(ws.max_column, 0, -1):  # Começa da última coluna e vai até a primeira
    # Verificar se todos os valores da coluna (exceto o cabeçalho) são 0 ou vazios
    all_zero_or_empty = True
    for row in range(2, ws.max_row + 1):  # Começa da linha 2 para ignorar o cabeçalho
        cell_value = ws.cell(row=row, column=col).value
        if cell_value not in (0, None):  # Verifica se o valor não é 0 nem vazio (None)
            all_zero_or_empty = False
            break
    # Exclui a coluna se todos os valores forem 0 ou vazios
    if all_zero_or_empty:
        ws.delete_cols(col)
# Salva as alterações feitas na planilha     
wb.save(caminho_completo)


#                                     ENVIAR EMAIL
imagem_path = 'M:\TI\ROBOS\ROBOS PYTHON\MULTAS\Robo colar comprovantes\logo.png
de = "robo@dukar.com.br"
para = "informatica7@dukar.com.br"
assunto = "Aprovação de CPs"
corpo = """ <p>Prezados,</p><p>Segue em anexo a planilha de pedido de compras.</p><p><br></p><p>Atenciosamente,</p><table width="501" style="width: 375.65pt; border-collapse: collapse; mso-yfti-tbllook: 1184; mso-padding-alt: 0cm 0cm 0cm 0cm;" border="0" cellspacing="0" cellpadding="0"><tbody><tr style="height: 4pt; mso-yfti-irow: 0; mso-yfti-firstrow: yes; mso-yfti-lastrow: yes;"><td width="188" style="padding: 0cm 5.4pt; border: rgb(0, 0, 0); border-image: none; width: 140.85pt; height: 4pt; background-color: transparent;"><p style="margin: 0cm; mso-line-height-alt: 4.0pt;"><span style="color: rgb(0, 112, 192); mso-fareast-language: PT-BR;"><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAKoAAABKCAYAAAArMs8AAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsQAAA7EAZUrDhsAAB0rSURBVHhe7V0JdBvVuQ5bF6ClLYE0aYJjS5oZ2QkpL+xtE7pQAtksjRSWBMJWE2Jb20iOkwCGUMrjsLehNNCSJra12SEQUvYApZz20df2tbRAS7fHI/S1aUtKG7ZAovd9V1eKbM+MJFt2xDv6zvmPpJn//nf77n/XGY2roYZqxlFp/+HOhOda+bOGGqoTDWn/Ea64npQ/a6ihOiGImvB2y581VCkO0DY0H+na0OxW+vQzlbjeqib1a9SUfquS9N6pJPRvqQnvWvy+QUl4O9Wkb7Ez6TnZccfCKZNTvg9LG+9r1IhapXD3e+vUtG8xiHiXmtB/AGL+N0j4jnvTokzjlnMzTQ8uzsr3zsvKg5CHsr/d952d0VK+DML8HeGfU5Oe+9SUf5Uj1XzK5NTJ70vi1ohaRVATC6cqieYIPOT34B3/3rjlnEzj1vNAzHMy7s1nZ9z9izJanz+jpX2CiJaS9mfc0BOkvg82HgCxt56LsL6MktJ/CdvfcMQ9c8Z1jTtQRl31qBG1CqDEPV8AgdaDQK+RUCQWSVaUkGUKyUviMg735kUZNeX7Dwwlwq7U3E/JpFQtakTdj3DGPadrKX2L8Hwkz70gpwnBRkVI2vtJ2vMyWr9vu5rwr6pbe9YnZdKqDjWi7geo3QtUtd8X1/r9We/G7tyMTGMkHFaQsBgTv+JM6cvGdfk/IJNaNagRdSyR9h+EyU0HvNlOEsMNopoRx1Y4HMiJ2f0RSCMmYW5M1JQ+/1OO3uaZMtVVgRpRxwiOjfOd8FiPiPEnPJgZUfJCErJrhh71c5Mh8Z0TLHTZ/GwacD03rh2hd0Z4TuDQiP6pxr2GTP5+R42oYwBnT/M8TFpe5dKRJZFIEHg0QTyQkETFZOf3alJ/SE367oKscYE4mHi1Kgnv5WrS0841UyXpu1FNeBNKQv8R9HeyEeTILYhrFlcJkrOD773O2+d8VGZlv6FG1FEGSYWx6B56wcFkoHACJSZSIAaI9pIrqX/blfRdxgX7uvULPybNlARlo6/eFffMBaGvEEtcaf0ftC2Wt4YzDkYYrtUqKf2HXDqT0ewX1Ig6inD2+q8kQcVsPjmQBOzWxYw75fsrxq1xkHRBY9p/uAxaESip+fVoKBHYf4rDCXr04YyLGQ5kfaluQ7Nbmh5z1Ig6SlDi3i4xbuQifUGlu+9Flwovpya921H5146Vp3IlvbPUfl8KDeId4cHL9LAMo6R9f3RtOGu/kLVG1FGAo8cTddOTFowRSQyxFNXv/xtIumZ/rVlySAGybhUNxmI4YiUcs2I48YLjnoVTpLkxQ42oFYaze4FP2+TfM4Ckm0FaTJTYxfNgiVTdr3B2e3xqn++3YvhRhncVw4Ck/tTEdfMOlabGBDWiVhCO+PwmrQ8TmPvgqTgm5bhQEMH3qjOlL5FqVYPx3Z6JaDzfyY6jzx4yjrYSHoBRej13STNjghpRK4QJG08/DDPtZwUxUeGctIjlqD7fQ87vznFItaoECNCi9ft2iWUxE2IOkb5Foodw9Sy8UJoYddSIWiG4erxrGklMVCS9EytSTXlvHdc1+2Bxv9fbIBRHiLq0/5P1d8+fIH9WDI7EwtO4hcpNhCHENBH2Ghiv/rVujCaDNaJWAPVJzwnwnOKsqNh1whjVFfcE5O1xDZs8R6tJ/UWt33+fq6fZq3771I/IW2VDSXkv0TafvV1Nea6BR/u0vFwROOJ6k9qv/1Yu9BcVMV5N6H0y+KiiRtSRIjPuACXh3Sp2lEjS/kXvuFL6ufKugNI7b7ya9L0y/dHzsysBm/wvYix77XAmVs6056JpD2cPSGtp/S00gG3OhPc8xiFVRgRxYCbt+70ga7ExK4Y3zA9Pgcngo4YaUUcIJTF/Ibt5sdTT73/b2btQl7fyyBJV/4MgstgqPUcQDeR9AxOuh0H080vdhVLT3gtFXJiocY2WXTW3XtWE/rKS9t2uwLtL1WHDlZw7Aw1ph5gUmhG0QMSSVVz/cWN6dE9c1Yg6Qig9nsemPbKEXnK3Em9eJC8PwACiFlQyvZEgGiYxakr/nSvpXetINZ8ig5mikKh5WyA/hxzCy6b099SUbxufq2rY6DlaBisbrviCucxT4TKbqTBupMfRu7BZBh0V1Ig6Ajh6Fnzeff/Ze9zwPM6e5mXy8hBYEXVfZcstVdGd+/aoSe/jzj592fhNnonSRB6mRC0QrjaIZ6q4+9Wnb1fS3jsdSe8Zw3nsxNHbvJo7UpYHaaSIXau452EOg2TQiqNG1BFAjXsT059cmnH22r8YoShRCyRHNLHNmtL/F6Rd6+jTT+NZVtqCxzzXjqiFktuu5VkDJan/EN4/7Ew0l75UhjgxpHis2OSKu25q2r8XxLbtDUaCGlGHCaX33PGo/HcwvrxfXrJEOUQtlBzROGlBd/6sa6PnAgwPVorF+RKImhcQicMLeXzwHyB/Nx/uK2V3SYnPPxbj1TeLDQF4htWV1K+XwSqOGlGHCVfa2wKP9zJJKC9ZYrhEzQk9liCa2PHS3zbTKUlAbhIuS34Qr8/3HPKxsqHX65JJNQV6jjUijJlNKfTyaAC/aOxqHJVJVY2ow0HX7IPhSeOlLsuMlKh5IdHK2Je3FZIf5JIEfB0e+0FnXL/YbALmTPmOQu/xZ7FGbGYLwnS5kb6mxFmnymAVRY2owwAX8F3dnrnyZ1FUjKiVFi6VgaxNDy8WHhse8aE6i00EpP8aTtBM7Ugh6V0J/yoZpKKoEXU4WNdyiPxWEipCVNn9C69W4uERK8l1/9ltXv3nStLb5YwvOE4m1xQ82oeJ1U4xPjaxScmuqXpG5UVmNaKOAYZNVHb1IIZYttrkzygpzxOuhGcLJ1e8ZxrGQsQ4l96TGw0p32tKSu+FB9XrumZ/SCazKEDUzXYrALSvJPSf5lYoKokaUccA5RJVLFFxI4CkSOuvqmnvd3PjYRBsEQlRElELic7X+SS9P9H6vZ1KylcvElYmXPFm26UxxqUmfX+pj1f+0EyNqGOAkohKUuE+jwlCdw+I9QyfOJ02qNKLLfhT8qsED4gdr79oKb1HSSz84sx1M8sasgwGT26pCe8Oqze6sIFpSf09V2/ziTJIxVAj6hjAjqg578ndLXz/E18fycdFEMx0l8eSqAVEx/e9XHdVkr5LGzbMO0YGrQiQvm1iPbYw7sI0IG3ORPM8qV4x1Ig6BhhCVFQoD5JwQqOm9T3ozh929YCA8FgyiCUGE3XfMAHes0/fjq73FiXt/2zuHGylgXx8026cyobCQzZSvWKoEXUMkCOqOIBCAdHUtO83Stp7nfSeJUMQld36/dJ7pkD0pPcRZ9p30WiMDQfDlWiO2S3+izQlvC1SvWKoEXUMIBfM+RaTN9Elb+J51eE+ww+CXzbt0fM59vwNSHqTK77gJHlrTODs9Sym9zYjKSU7xvZdJtUrhhpRxwB1mxd+zBXXA44+vUleGjZcibmfo8dSnxn+UwIjgZr0niH2/TFhsyIqhzFSvWKoEbVyGLUjbtUEd0L/HAlptZXLk1983EaqVwx5oiqa1uV2u/s0TUsNFpWiqgnV7d4AvTvx/d9dLlfAqWlfVhSl3ul0flDaKwmIZ5aMKz04Lgru9bs0bcg4x1VCOKQvItXzwLVL3Y2N/WZhtMbGNPKXdDgcTqmeR0NDwzTcT0LPND6WC+2iPG6GuuVhDJTRCY2NjQxjaofCfEHvNr+/+GI5yvujiPMuzaK+Bgnzl2LdQW52KsrlqLsZ5dZZDmpCP03M7k2IKq6lfXuVdPNnpXrFsI+obvd/Tj/22EzTtGklSWNTUwaVlAEJ/omC+Bk+v84CkHZtgUq5xC4u3gNRe6R6HiRcsXBIy0NSPQ/kbcOxM2aYhpk2fXrGjbyAJJ+R6jkcqKjqY8XKhOHZaGUYU4Ag9xazQ2GZwtaXZDBL1NfXT0A+32LcZnbshHEg7LuQ/0JcsQkTJhwmzZYENe5ZILZxTQ5Tc30Vk8ad7h5vnVSvGAqJ+iQzgVZalpCsDMdCwO9dLlVdi9Z6lLRvChD1fLu4eA8kGfKCA5D3gqLhNK1fqueBa3fK9A0Rph+VtgeVNmBSgrg8JDDvm4Wj0CY8W1oGMQV7HNjfZWcnJ42wB4+3Xga1BDz90bD551JsmklhnaER/Qh5L/khQ67NWs36xSPUSf2F0Xh+asRELRQWgMi8pv0KXanlae9qJ6roWjXt+WJx4XMHdG1PzIOol9vZKRSZ/lcmT578CRncFCMlaqHI+noRQ4Qhj72YAZNCy1NUHJ8qSe8DUrWiqChRcyIzv1NpaBjcnQpUO1HZJVrp54TeFnq2s9uZ48YdgnL9UTnlSl30SrZ2K0lUCvOK8l4nzdtCSeq9Vgv+2WN+nhukakUxKkSlCFua9urUqVOHjFeqmajoyqfi+g47EjAe6G2Fuu0Dc7A5i/rlEIrpBFG3IbjlKkKpRGU6KbRppyvLYAe8v+3TCjwroCa9PxM7Y4OJKp+EVVOeBVK9oiiLqMwwhXqlFD51UaEpxDOg0KuUqOIgBSprnZUuRcb/mtkqwWCAcDfb2TITpgd5fwNktHw8pBhRZZ5eQNkvhlxIb4mx6Ft2dSbKTlFsH3lu2OiZhln9W4Pf+UoRJE14t5NQUr2iKImoMuO7kek7IF/D94cgfxf6NpnPFQzGcbNlfAJVSlQ38jYd+Xm7hApdLk1bAmQ6Anb/xyy9yD8b8A6Q5+3B9yicVDlcrtXS1BAUIyrjhP3HpboA0rzCruxYPuBAh1Q3hSPhXc2H+AaTlJLt9r3fkaoVRzlE3dXkcORf4NrQ2HgMMn+bXaVSWAAu6MlgAlVI1N0g6gzEeS9JYqZHEZWpqo9Js7ZwKIplHhknyuRRxGs6xGA43Htx4syZpk+JlkjUp6S6AHqAJtzbaxVG1JPLdYVUH4LZT84+mOdZzbv97Izf0esdtZdQlEPUNzDmnCbD5YEC2WAVjiIL/VczZ+47C1mFRN2JRvd1/H538P2c5PTQOzRKs7ZAuWy2SittgRSrkK5fmxGHHpfiUNXPS3MDUApRUVdPSHUBp6bNs9KnMAzydpFUHwJnb/NsrX/Re2YL/eKQTcL7/ISNS8paky0HIyZqg6oej3u7WbA2YXfVYZIig1QVUSlIHxfA9xSrSBAnKk3aAkRywd6/zOzJ8ngdOsfg83ErDy493B3S5ACU6FGf5TiaQpLi96+syk6m6Q3EZ/lqTCWh32+5LIXrjqSnXaqOCkZM1ClTpkzCvb9YFZoM+x48Vv7lXdVG1GJCu8gDZ+IlAZOoNVbxybz9kHpIV1cRT//aeJP1zWJEpYgyz+4a/ou/7XS5w4U0xaX5IXB2e07W+nzviRP8g0mKLl9J6n8ayXuuSsGIicqCtCs0GfY9V2Nj/vGE9xtRuWaKPDw/efLkUv6F+RB4r59bpZMeFI32diry7AImb3vN9Ci0gUnVBcJqAUohKoX37XR4j+WCSd0vG+HhpfmByL7WZ5vp2inHpvCmfGRGao8aRkzUekU5FjqWM2UZ9s3C5RxU5OJihIPO3VI9DxB1qW04klHThrxYtgSiWpIlJwwPghXddUH3OQt5tiUfdDzU5Q4U8vmyVdlRFw12SJylEtVOEJ7DnZ0om3tQN5b/cuLsbr5YPHpisrdP8ioJ73MTt4z+H0+MmKjI6G225ME9VMZLGKjnT+s4VdVnV8gkHMIkpHoeIErILi5Jpl6pnocdUWU6XichBt8rFOoxbqei2C7hIK5eu7hQzq9+wunM/2Uj9a3GqTJtb9fX10+X6gIjJaroIdBApk6dqkqTpuB/D6gpfYf5s16Lsm/W7vUWPURTCZRD1F3osvO7TCisI1BpV+K+7SREkGfQ9hwK6TS7MILcmvYCVAfs/ICE3y2BqEO2Aksg6lvwKqdA70kr0lCkLnsH03MM4lSTqtqP193uv9EzQ7a6FOV+xGk688+JbBxdMgqBcrp+s+sU3OO8ISxNDkX2n7Afafqe+QSKj2Ar8YXfkNqjjnKI+ja6rKshbax4/P4t9YsUhhAUyBwZnwAqU8V924V13sOk5FbqchuWSyfw6K/bhRFk1LSVMpo8ihEVednDeJC3k3Btt10cuUbEhirN58E9eruGlBPqCEGa7OKiUA8N/aezZ+97WK8YUUWeVPV55PuPVumR+d7dMPR4o4CS8Nwk/jxDPkSYlyRP8Z/L5agf859gpPqooySi5oSZow4rvZguRXg4VX20cA2VAOl4Oul3pcQHvV2Qf/A7xUyvUJqamoa8pKsUoqIxiVUJfNpuoVJEvhRl8C7MQSDHM8XCli0y3yiz/AvZihFVlKumbYYnXmJXZtRD2TxXV1c34LXsznjzMnb3Q2b5IKlY8O/3/7lh4+lDhoGjibKIWo7QFgrTcoEcE6ONpVQqC9qusHMiCh1eBIU+5BU1pRCV3pS648eP5yqG6dZnoTAcyJp/LBj5PA7h3i0lreWKJFS+YZRIVO6gHYx8/cAuLywX9ARrs5bHjXN0Ny8Sr0Q3eceUGKv2+9/GLH9ADzkWGBWiSjtvofLmy3iGgKsFGK+9WYk4WWHwpCTO5dL8AJRDVIJLQsXSJcO9BtIIzwJvusZufEuhTSux2jCh8D4bT264UQpRkR6xhYrx9KkgrWUDktf3OifVz3M8vWSO2u97RxCy8EVs+C5O9d/re8/Zu2Ax7Y41KkpUZloSYjtIc6aMwxIc74pKsijEUoVxwptywdr02F25RAUOgpd50CpMTiSBHuOmB+L/TbHyQzqeh/73IT8okKfRYJ/B55t25SDJJ86plkNUAnlJ2+XFrbkzapP7NfXuhbvc3H0aRFLxryv9/ncb4s1LpckxR0WImiMoP0mYwjXTYgBBYrCxdzhxMwzjRaUkjlRVy8eHh0HU3CGON+zIQxF5VpSfwMZus/sUGQdXFiwfl0Yat9iVAdOPshUHYrRJk44sh6j1mqZAf6ddXtwOLePWTxJvDSx8FFr8k1+f/5/OuMcnze0X5ImKVv0st9KYyWLCQqPwu8zoH9G99KDChjV2wRBhHirqJ8I+7Fo9q8RrlFw6UPivIE4+dWr7qDJsr+fDdblwhcJ80Ca7SKmeB2yLLU6mySxsTnJlYSVii1LTfj0bY0ZpegjQUFbblT/jYBlw7ZMP5CHvu6zipR3Ux4+laQF41RsK7btdWkZrUPdJvZLRPuXMaG2z0cVnn9vnn/Oqad9Lrh7vLGlmv2GfR9W0mzBOegIF8NhgQUt+FPcfRsU9gN9pfN6DjN+Mz+Uo4BOPmT7949LesAGyfhBeYDEKeAviexnxiGWifMFmibsX13fQWyDuFewCZXBbQD+I8E/l8lMoiG+byJ+iaFI9j4kTJx6K672I90mzsFK2wf7TsPE0vj8+6J4QETfSIM2aAg1lJtMCfVMbFK2x8WnonQH1g/A7jXSZ1hfq8UnEx0e485ioKONRh/2aqj2uTXH1q5+Z/gdt/gkZbe7x++TLMzOa58SMu0fPNJKkfb7Nzm7PZGlivyJP1GrCJHRtKOzjQJKF3N9Ho7iQu1lcekLDKGW//f8zRvySXHjJxdqmRdvFG0+4TiqFv8Uaadq3U0n6QlK9KlCVRK1hdDDlvnMmqZt83ZzBZ5/NJ0ExHuUbAUFQ8ar0Tf5Nzu4FJZ25HUsMIGrIMJYaxoobQqFo56pVq0p6fJZobW09MhgMikXjQCBwTDgcndPeHjmuq6vL9uE33g9Fou3RaOc14bBxRSAQs/0bmdFGR0f75M7OTtNhTHu70RgOh00PMre2dh7JPMuflkA+T4l2dF4biXZ2BcIdbe3t7UXfVgKdyRdfHBvxe6bgIZdrfb7t2bdPZydL3K8veLvg9/lyCaledcgTNRyMXgCyXI2KcrYFDW8oZIj/9QSZDg6HY2dFIpEvihBAMBhTg5HY2SCnOAwdihi3hqNR0VW0h8NzgpHoFqM9tiKTyYhJTnskclwoElmMQh/wYgqSO2IYGzo6Vp0YCEc9wXB0Y66BxGKxfwtHO84DOXLPuB+ANM0NBo0zmSak80g0igl8BU6bYYhXfTMsCYOG9qVAwJiJNDZDV7wMAY1pUjAYOXf58uXivALyMx62ZjGfuUbG+ANhQzwPFQqFGphmpkPcC0b8obDxVX4vBMJ/IRCKrAoGo2Lf29/l/wDS4GkLh4e8dTkQiq2CI7hq5cqVGuJ0M+1MF/WRHoU6hgF74aiee7VPKBTrRnmLf6tua4touLckEFgp1m6ZN5FX1A++m76Ln2+aFsf0QEYuM5Gk/E7CYgy6W0vpW5T+RXyEpKrfnZUnKgqap8kPQMV8JBCJzYNHFLPgtrboZajQ64Jh40YQ+Ry/vwsVYaxvC8euCIU6xJIFvOLtqFRxwpvXUKlnh8Kx9atXr54Cok1HZd4cjEQuho3rUaD51y3SExvR6HXyJyoyeiULPhqNflqQP2xcEgzHbqDnCRmdLai0VajUECoZDSVyRgSNZYlhHBY2YuJZn0gkuhqe7yro3AlbXYhvLQm2omXFEUjDXSBBgGRcvjz6Sfz2BUORBMKAI7GvMXwoEutmPPzO9EKnBfpbqI98zGmD1+e9HAKB6BKktwtyDcj0dV4jmUlcxsdGJRQlwuGOUCBkfAdxXtgW7hBEbjdiK6Dbizw72tpizbQVCHS0oUzF0wQoh26WaUtLyyHt7dFWNKBlyNOWlhUiTz6UTwr2Lke3NOBFGM64fpzap29w3+vP8E+F6Tn5pj1ui6pJ3y+UtO/GSvwT9VhhH1HD0dvaOzomo0AORaF9JRgyxAFkVN634J2MtrYVsVAo0sqWjgK/iEQIhDqX0WtSH570ZBYmyHQpWv15qJB4IBZzBcLhS0JGTGw1Isy1qPD8Gmsnu0zDyP8tIQmP+zNgvw12xA4I0nULPWY4El2HeMQbmdvbVx4Fci1gRcGtHBg0Ylfzelt4RYw6uL4Idk7MelejpRWNBb3CM6jg5mAoegd7AvQA58HrXRYO3/zhcKRDnLiCztWGceUJCPthds3BNvQywdjWSER4wDPaw8ZV1MsBtr8FlmuG0XU0Gvp12bKJPo3vX8G9r6K8BuzigIRoFdHVyN/01tbYJF5DnDGUnRhS4P7Xli83hMcHyXvYc6BMOkOdoQZ6fdRFEPaXoG6eQqOrF15feHNjjXH5lYJ0TQn9VCWhJ0jMadsuEB5UTeqv4NoTakq/3pX0zqpbX/o/oVQL8kRF93wqCvebrGQUWGswGF5DBZIShdHOSsW9L2Qr0VhOMsKLrBWFGTKWcqwZ7Or6GIh1N37PBSHuge50ej/Yuxbj30Ug6jUo4PyJG3TDh4Oo60NGx1LodMCDXgd7B4Kkx6MCr896ZuOrLevWHUIvhOvBQHiFAa//OdhppJdHg4DXi91Ie0hD63LDqDOMDhL1ZKRjPtJ+KSs5GI7cgkajt4djVy1f3nU4e4coCBXp6hqPRnCLCI+xOcK0wPbRyEcK+mfhd5JxBTD0gafspF4OII2OngRhoh0g5038d2WSmWklAdFoBhwHhH4IXnjA//mzbKEnznTCizdnhxGRi9EAxFsJkacr21EnrRgahMLReDsaH4j6AIkKlc8G4SQikY7Vyzoil0++b95FIOTDasr3CMj5DVRsS0P8zOP5xmvaej9jwGQKma9D5k/noD83RsJI6ECQ9USOuWaDlLzCQgoaxix6Nv5G1/zRQDR6LL/TWwWDnSdRB6QTJ7/Z8kGcz8PjDjkaRxKgUmczDpJUXkaYTngR4zSSSl4aB8d+Ase78ieGJfS+4ZNhQ6ynIu3jYeMDJOaSJcZhTBev8R6HGUwD88jfnDTxGhsaiC+9W/gTSKvYPWL8zDPzwV6GjSoXTyEQZgbzi3GjeC057UWjKz8DQoryKATH6MFg14CTSkwfh1vyJyZc0U/j2qm5suA4nBM5fufcAL3bCRgCOJYuXfohHv1juTHMzHUthx6zde7HR+M/nqoBA4haQw3VChLVmfBaPoBYQw1VAb4K3pXw3vp/y6t70lo9DC0AAAAASUVORK5CYII="></span></p></td><td width="188" style="padding: 0cm 5.4pt; border: rgb(0, 0, 0); border-image: none; width: 141pt; height: 4pt; background-color: transparent;"><p style="margin: 0cm; line-height: 105%;"><b><span style="color: rgb(64, 64, 64); line-height: 105%; font-size: 8pt; mso-fareast-language: PT-BR;"><font face="Verdana">Robô</font> <font face="Verdana">Dukar</font><b><span style="color: rgb(64, 64, 64); line-height: 105%; font-size: 8pt; mso-fareast-language: PT-BR;"><br></span></b></span></b></p><font face="Times New Roman"><p style="margin: 0cm; line-height: 105%;"><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>(31) 2555 7850<span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'><br></span></span></p></font><p style="margin: 0cm; line-height: 105%;"><font face="Times New Roman">  <span style="mso-fareast-language: PT-BR;"><a href="mailto:informatica1@dukar.com.br"><span style="line-height: 105%; font-size: 8pt;"><font face="Verdana">informatica1@dukar.com.br</font></span></a></span></font></p><p style="margin: 0cm; line-height: 105%;"><font face="Times New Roman"><span style="mso-fareast-language: PT-BR;"><strong><font color="#404040" face="Verdana" size="1">Email gerado automaticamente.</font></strong></span></font></p><font face="Times New Roman"></font></td><td width="125" valign="top" style="padding: 0cm; border: rgb(0, 0, 0); border-image: none; width: 93.8pt; height: 4pt; background-color: transparent;"><font face="Times New Roman"></font><p style="margin: 0cm 0cm 0cm 1.7pt; line-height: 105%; text-indent: 7.1pt;"><b><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>&nbsp;</span></b></p><font face="Times New Roman"></font><p style="margin: 0cm 0cm 0cm 1.7pt; line-height: 105%; text-indent: 7.1pt;"><b><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>&nbsp;&nbsp;&nbsp;<img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAADEAAAAxCAYAAABznEEcAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsQAAA7EAZUrDhsAAA6/SURBVGhDzVoJdFRVEk0URwUVHUUMSELS/X//LOASXCYqcQENWbv/kqT3JSEHEEacwVFxJDi4IIpHFFQYICFJ70kM+06M4MgIIzrDQQY4DqggyKLEDUax59br14EsIox4zD3nn9+vfr33qupV1av3f8eda6TN1C4R/MZUXcB0Y2q9nEQ0wSsPE/3yJqmxZLXoN04S6+TbGHN3g+iVJZ3f+JhUrzaLAXmfGFAOSUH1gOgztYhetVwIyLlSg9aSscoWSV9hjRiCSqO0wHgl7949kBbSEiVY3xBUDbr5rgEpaNOKQIm79T5llBDIHUp8hoCqSSGtGavyPZRdCVK8wVtUIIWV/LiKzAuIp1sCCvWGG42A9adDgXoouszgk9/E6rQKAZNDqoXyYe3z1MXmTaQQ79Z9kDm74gLBZzJB4PVSSP3aEFC+h5u1QpFdYkh5W/DL84gPcZObtsQcSV9miQi1psmsc3eEPqj2EWvUZLjeNQmz83tSW6iXU/Qzci5kz2vgVk0ln6YttUT0PtNwkOKJ3q2R7DeKWIm/SmF1uxhU6KrT1RlnpTYVfwTX8+tqCvSIqWkps4f15l26JwbWyUlCUJkMV3sH17dSvRZBwO/V++X1Sf6igQa/8rFYr23WVxt1vEv3BrkZ4iXb4NUmIqNNT1iU31P0m6YMXu+mlLwlpcZ0NWc9M+TnV/R0uVxZTqfTaLO5CqxW5502m+06h8ORbrG4UyorK8/jrOcUSd58iTbKpJm511A70Wu5AoodHLTOEYFrvcyYfgoQtJfD4XoU13Kn0/2iw+E22WyejNLS0n54lmW321OhVBaUMePqz7v9X7BYnNdjrAlkME5C8BemCV7T+4Z69SA2x9eR1Ybp6kwT0pdjgwzIu6gi4KxdA1Yf6HS6/m63Oxd4PJ4MTmbAZI9CqX9Yre40alut1mttNuefysrKfssYzhIWiycJhtpvt7snZ2ZWtNvkWCwE1Y2DWpws5Yo+eT6utxAfX+jn5fThbF0Dln0GA8/kzTZAuXGgf2W1uu7gJAbQH7fbXRbePGNomnY+VvhNGKyWkzqB0i7KlTHYHD+IKqLsQv21kQJc8BXewtk6A+5yi9nslHiTAXFxByY7CoHv4aQ2gL4Yyo3lzU7Aal5aWZndgzdjiIcCD6LfHk0bc3rXIETi4qVGZaToV3YgHc8SA6abkL12UUnDOU4PCH45BN2MCcs5qQ0I9DzQW91udx+Xa/QA/H4C118QM0NKS8v6QtBZDodzEVxOIX6s8r1w0/G47BjzEMbOIbrHY8uI9nNOQYxkEi2GzM3Y7YPGm+l3ijdPSPEabxe8xpvTFpspPhoYU1fAUv8mAu3pN+IAgjk6uRdoN2Pir3GfQBPj9zZcr4L/eauVhHX5cM2gjGax2PMh4GQ8XxPNbu4GtMPRcdwq4mwH2i9CqRcsFsfQ7OxTVg5yoBp+2dCoLkmcMzyZSEKdPDS1oQQpVz2R7FMGM74YyKLkSpjoalLCavUYIEiVqqoXcxYGp9Njdbncx2hiKNEfQhyAgDOhQAH4cyngsRpOPFM1zZ6IZy/B8jspq0Ho0eBHMNsTwYIVcZ/A80ooWoQVywFdbKcEQG6T2lAcSV1YclT0GYt1voJ0KaCcoNXQe01Wzsb8/iYMcJ+qnsw0aGuge3iT4iUDwix0u8tbMfl4opEAuHZbLJZBpEBpqX0I7kVwo1V2u2MNDJAMwT8HzzN2u4ee+fBsbXGx6ybQV4H+Nilns5UVYCWup4Bnk3WA6DVOGrTWEZEaio/rEezIUh+mo74SfEYzY7DZ3LfE/LYjKDBh4Szo4nO7yz7GPRxLsQQIMASCPAue4ZQyKVVivGooOQtCSfn5+T1h8UkQ/j74RjxoeWg/BqWvoCSC1XoSY2TTivIhu0Zzdg9kpjmkCO7vIbB3okQ5pptfNIA9t9kcyP/MCr/BdQmsoscKYBU8z8LyLbhvgQIvkSVZh18RBr+8AKX8dziLfIGT5COGuVmXDqwtNCDLuG4ka2KZkU1cc+HvVfhdDeGRMVxFUPByPsavBhyaEqQgK83jUHq8iVS7nP32m6r1AeVJ+k2Ip52XBO4YWN0Bhirs4AH5E31twa1CdV5/saYAZ5L8RFS8J0B/g7N1byR6864gN5Lq1c8Eb14K0agYjL5wUD9mTN0eIe18lOJvD2p2Rgw+ZTVTyidvowyFfWQ75/pRxFOG4r/bYcyYMdfY7WWpZrP5Kk5qB+wXacg8ibz5syH4lYnM8gHlqxRvYSmUOJyx0kbl+YucpTOo9EYuX4v0Z+MkBrQLEfhNCPqFSJd/hKDyqcFPKRjPlyJJzMDz8arq7tMx/7tcdg1pvZQyXjSFuzzIhBujqbhrCMG8/nCdAzhb0MlvDm2AKNe/MQSRnboCrHgZBn0P1yuUeolmNrsEj6dsBSbaSjsvhL/SbLbfjjLjgRgPJQUIvx7CT6U2AWXJbVRL8SYZYTDG3Yz++UjvDpTiT6PPq7gm0G7P2dohubqAlRf6gFyJQD6BYnB3ar3WmuIz3c0YOoIEgmGaMdF6TqJC8B6UG0ewW2MjO2l1bGpUTrfwJgR2p6C9n1yN2rAwDlHOkTElqYSBsBtiO/6ZQF+bcxlcpoWOrjpv/u+QXg9DiRZ9bWHaVa/dl6DzmQo5axQ0GawdxOZ2HDuqnmi8aPsWkz/BmDig1LgxY8ZG6M5JcTk5ORdCiX+Bt5yEp0NTTGmM3RvCN+H5N7D+MCp16HzSsTbrCMmnmDPW2COGWuMDfWtsvXD29lwbVC+WFhhFqbH4I5TmbftEHO0TtOGRYJgkZimUCm74tzPA2wxQSoWgH0LZI3CBBE5mCmMVd6LPUbhNMSczoD0CCq+lQpHcNboqVA07p5NrcrZOQJkxj5cbm+NCcSy2dD4l3RDW9mWsYfQSxogKNg1Wux/CNePaAWF6EZ3uaG+x28tFalN9BL4K+PJYCErHWB/RKQvBsjUQcqPTWYad373HZht1Rm8mYIxmjHmyGu0AZKPllEpRK/0gevMlfTCnDwJ8N532cI8keWWJBE3GIDlkRQh2BL9LeX8GtF0I5j8QnbIILoECErxfW60jEyD0bSDvg+CBWCxAsI24nmcDAHCl8+FWF/FmGygDwr2OYHyZkzoBmWhx+hIzlNB+EGuL7kE8LEpbYomevfGMMZEf0x010zQM+O/Kysq2siMrK4vtEVj2/qWlnn70m84ZUOC/UIQyVCL6fIkVeIiexUCHf/C8Ff1tvwFC5lJMsIcncR7mXAPlP7Jax13GaW0gl0mqyr6I0mnGKju5zT9Rev85takkQkoZwupRoa7oes4ejQdMehDuYeckEuT2kpKSdudYCJTqcpV/BsHZAR/WnksXe3gKINwc8KyE79+BMR/TtIp2CtAKIJ7WYAWPUVxwchtw2HHjzLCWdmrJrzxPr2tQ7K2Dax1mQR5SdwvewmGcPQqHw1OISXdmZUV354qKit44F4xmDzlgMQcmPoRrHrlHlOaooiMlY+CA2+RAsW/JReB01VCcxROBVg7x9zhS9X6MsxW87c7UBL23SKG34xB4FbVRVnihSCuUOCLVF283BOUHUl7p4k0gJh2LCVfzZiyfT4PgOE566HXNcrfbswX3di8MysvLb0Ws/J4OVvQbQj0Eng3YABkflHwSxgmj78u0MpS5kMI3ejzl4yids0FOgX5B4Q3w/VZyHymo1IMULwTkqTiiviDUmkwD5hb2ExeWFAkBUymdv6O9OGinpOA+9dUkLJaCSSfCzabAqlosdjoCCeFq9DViZx4JJRQ6tfFHcTQe6isNQj+MZ6Mw5nX8UZcQfQhiZKL0ZXCfQPS7BYF9EvDLo6Wm4q2prxfDtYo6vT76VYGdmK0IvZqhNJraWBxJW1gakXym4Wy3DuEEF1Y/JeXYfuE1TWEduwOues2UIPjkR2NWheWforSZtsgcAb0uGYGLQ8/7FB9Eo+BGyVEbVxn3i7zIPivoZ8C6AbUSwn0CoT40bMi6lL7pGULa7rTFpScMftMG0GdiR/6KuRb2BFodKDkjtmqdIAYLkqUGU0I7DZuze9DgfWuG94JjnzPNUxaMGILy+V3yeToLCCHZRXTUP1Wog44jA+0Rg8oBWhESHgeh74Sguoy+57EBfgx6v/FewWd6S4AF0OlvuN5BevsA9z2478b9PZS/TbDQs/BNjb0mOY1i2TAAfZ4SllhS6HMW5Xqi0+dhKaztpY2KBEQQ/4c2M32t6VoDvWP1m57GZjZVDKmVQlAeK/oUm1BjOm0iOIlI5XlQ4gks1xEIvd8QUL+jZcyAD5K1MlbbI3QspJ0SSrTofcZRtEq8N4MuZMwUg9rDUDaE/u9KDcVfQJhNMX8fWJebJIWV7TQuvbnDWeAgaNH9pTK7RyIO/un+3CyxzjgCijqhjJM9OxuQryGFVWPyVgi6VgwrftTwVbBQLX0MNIRLJgoh7eQWHwXl8KFSQPUa6pXj7PMtFKcvOjjQb6KPhsREL4Cxw7LX8/TpV6zXVqKs9hgCJg0Kz0ERtxHG+xzCR9JhNENY+RLnaA8Zl81yVsCyYzUa6RyLAQ9A+IU0ES0554jLnJ15gS4g3wWFHxED2mapQWMrlg7r0gUDfCMG5fkUvMSPtGnG6hyEYN9j5/0UxdsbmKMZQbyPViQDc7ELwmN1qJxoYi/Dfg7owwasMDd9KSbArpkGF4JC2yS/PLbvcwhy4vGZxrOsElIPQaC9UGi7FFJWwKpThZiLcLAvo9hdMeZ0JjzclSxO2QY7M5UWx0Dfg9WogbJ38W7nBkKd7GIpDy5CMUHZBFllK6zMDiD0ZxS4XJUUVo8g8A9LAfl93Fsg1HpYex2uMNLoU/S97dQ/oST7CvqSW+r8RXfqg8ZsqkL7/ZJ/UqFvY8gQ98Ot1sHqx8jfyWfpXzRJ0RI4HgkhDwG8LhU7LK0cuQUpzHhXRP9NQ5ZH0bYNFn/uFxX4p0BplawKN6qAUtP0fnlS/9qctjcTOi+yk095EG6B7KR8gOJtL9xrB9pLsQdMQx8bng9Oau58MPr5iIv7H+u+7vKhO9wlAAAAAElFTkSuQmCC"></span></b></p><font face="Times New Roman"></font></td></tr></tbody></table><p><b><span style='color: rgb(64, 64, 64); font-family: "Verdana",sans-serif; font-size: 3pt; mso-fareast-language: PT-BR;'><b><span style="color: rgb(64, 64, 64); font-size: 7.5pt; mso-fareast-language: PT-BR;"><font face="Times New Roman"><img width="495" height="3" style="width: 5.156in; height: 0.031in;" src="data:image/png;base64,R0lGODlh7wEDAHcAMSH+GlNvZnR3YXJlOiBNaWNyb3NvZnQgT2ZmaWNlACH5BAEAAAAALAAAAADtAQEAgQAAAKampqWlpQECAwIVhI6py+0Po5y02ouz3rz7D4bi+AEFADs=" border="0" v:shapes="Imagem_x0020_6"></font></span></b></span></b></p><p><b><span style="color: rgb(64, 64, 64); font-size: 3pt; mso-fareast-language: PT-BR;"><b><span style="color: rgb(64, 64, 64); font-size: 7.5pt; mso-fareast-language: PT-BR;"><font face="Verdana"><b><span style="font-size: 7.5pt; mso-fareast-language: PT-BR;">Rua Dr. Gordiano, 164 | Prado | Belo Horizonte / MG | 30411-080 | </span></b><span style="mso-fareast-language: PT-BR;"><a href="http://www.dukar.com.br"><b><span style="font-size: 7.5pt;"><font color"#0563c1"="">www.dukar.com.br</font></span></b></a></span></font></span></b></span></b></p> """

# Configuração  do servidor SMTP do Outlook
smtp_server = "mail.dukar.com.br"
smtp_port = 465
senha = "0za*s,}mx}e."

# Criação do objeto de mensagem
msg = MIMEMultipart()
msg['From'] = de
msg['To'] = para
msg['Subject'] = assunto

# Adicionando o corpo do e-mail
msg.attach(MIMEText(corpo, 'html'))

# Anexando a planilha de CPs para autorização
with open(caminho_completo, 'rb') as anexo:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(anexo.read())
    encoders.encode_base64(part)  # Codifica o arquivo em base64
    part.add_header(
        'Content-Disposition',
        f'attachment; filename="CPs para autorizar.xlsx"'
    )
    msg.attach(part)

# Conexão com o servidor SMTP e envio do e-mail
try:
    # Conectar ao servidor SMTP
    server = smtplib.SMTP_SSL(smtp_server, smtp_port)
    # server.starttls()  # Inicia a criptografia
    server.login(de, senha)  # Faz o login no servidor SMTP

    # Enviar o e-mail
    server.sendmail(de, para, msg.as_string())

    print("E-mail enviado com sucesso!")

    # Fechar a conexão
    server.quit()
except Exception as e:
    print(f"Ocorreu um erro ao enviar o e-mail: {e}")