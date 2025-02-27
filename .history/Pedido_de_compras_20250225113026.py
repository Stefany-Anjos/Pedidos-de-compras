import os
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# Função para verificar e excluir colunas com apenas valores 0
def excluir_coluna_se_zero(ws):
    # Iterar pelas colunas
    for col in range(1, ws.max_column + 1):
        todos_zero = True
        # Iterar pelas linhas da coluna
        for row in range(1, ws.max_row + 1):
            valor = aba.cell(row=row, column=col).value
            if valor != 0 and valor is not None:
                todos_zero = False
                break  # Se encontrar um valor maior que 0, não exclui a coluna
        
        # Excluir a coluna se todos os valores forem 0
        if todos_zero:
            aba.delete_cols(col)
            print(f"Coluna {col} excluída porque todos os valores eram 0.")
            break  # Interromper o loop após a exclusão de uma coluna

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()

    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select s.descricao, m.os, o.placa,c.nome_cliente, orc.honorarios, (orc.vistoria)validacao,
orc.taxa_detran, (orc.extras)Reembolso, m.observacao, orc.ipva_total, orc.prim_parcela, orc.seg_parcela,
orc.terc_parcela, orc.total_dpvat, orc.desconto, orc.placas, orc.lacre, orc.outros, orc.srpr, orc.lacracao,
orc.placas_esp, orc.licenciamento, orc.vlr_correio, orc.multas, orc.vlr_bo, orc.total, orc.txvistoriamovel,
m.solicitante
from os o
left join cliente c on c.codigo_cliente = o.cod_cliente
left join orcamento orc on orc.os = numero_os
left join cp_os on orc.os = num_os
left join contas_pagar cp on cod_cp = codigo_cp
left join manutencao m on m.os = numero_os
left join pendencias_os pdo on m.os = pdo.os
left join pendencias pd on pdo.codigo = pd.codigo
left join servicos s on o.cod_servico = s.codigo_servico

where pdo.codigo=120
and pdo.data_baixa='30-12-1899' """

    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp)
    # Resultado da consulta
    results = cursor.fetchall()

# Em caso de erro obtem e exibe o erro
except Exception as e:
    print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
dados = {
         'Descrição': [],
         'OS': [],
         'Placa': [],
         'Nome Cliente': [],
         'honorarios': [],
         'Validade': [],
         'Taxa Detran': [],
         'Reembolso': [],
         'Obrservação': [],
         'IPVA total': [],
         'Primeira parcela': [],
         'Segunda parcela': [],
         'Terceira parcela': [],
         'Total Dpvat': [],
         'Desconto': [],
         'Placas': [],
         'Lacre': [],
         'Outros': [],
         'Srpr': [],
         'Lacração': [],
         'Placas especial': [],
         'Licenciamento': [],
         'Valor correio': [],
         'Multas': [],
         'Valor B.O': [],
         'Total': [],
         'Taxa vostoria movel': [],
         'Solicitante': []
        }

# Crie um DataFrame com os dados
df = pd.DataFrame(dados)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)

wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)

# Atribuir os valores ao cabeçalho
for col, valor in enumerate(dados, start=1):
    ws.cell(row=1, column=col, value=valor)

# Adicionar cor de fundo e estilo ao cabeçalho (linha 1)
cor_fundo = PatternFill(start_color="C3ECBC", end_color="C3ECBC", fill_type="solid")  # Amarelo
font_estilo = Font(bold=True, color="000000")  # Texto em negrito e preto

# Aplicar as mudanças nas células do cabeçalho (linha 1)
for col in range(1, len(dados) + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = cor_fundo  # Cor de fundo
    celula.font = font_estilo  # Estilo da fonte (negrito e preto)
# Salva a planilha após as mudanças    
wb.save(caminho_completo)
# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(caminho_completo)
    ws = wb.active
    for i, row in enumerate(results, start=primeira_linha_vazia):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
wb.save(caminho_completo)
# Iterar pelas linhas, pegando o valor da coluna 9 (coluna 'I')
for linha in ws.iter_rows(min_row=1, min_col=9, max_col=9):
    for celula in linha:
         valor = celula.value
         # Verificar se o valor é uma string e contém o texto "CP R$"
         if isinstance(valor, str) and "CP" in valor:
            # Remover o texto "CP R$" e os valores seguintes
            valor_limpo = valor.replace("CP", "").strip()
            # Atualizar o valor da célula com o valor limpo
            celula.value = valor_limpo
