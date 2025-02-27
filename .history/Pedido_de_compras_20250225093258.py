import os
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
from openpyxl import load_workbook

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()

    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select s.descricao, m.os, o.placa,c.nome_cliente, orc.honorarios, (orc.vistoria)validacao,
orc.taxa_detran, (orc.extras)Reembolso, m.observacao, orc.ipva_total, orc.prim_parcela, orc.seg_parcela,
orc.terc_parcela, orc.total_dpvat, orc.desconto, orc.placas, orc.lacre, orc.outros, orc.srpr, orc.lacracao,
orc.placas_esp, orc.licenciamento, orc.vlr_correio, orc.multas, orc.vlr_bo, orc.total, orc.txvistoriamovel,
m.solicitante, o.cod_servico,
codigo_cp, (cp.ipva)ipva_cp, (cp.licenciamento)lic_cp, (cp.multas)multas_cp,
(cp.reembolso)reembolso_cp, cp.pagamento, cp.qtd_os
from os o
left join cliente c on c.codigo_cliente = o.cod_cliente
left join orcamento orc on orc.os = numero_os
left join cp_os on orc.os = num_os
left join contas_pagar cp on cod_cp = codigo_cp
left join manutencao m on m.os = numero_os
left join pendencias_os pdo on m.os = pdo.os
left join pendencias pd on pdo.codigo = pd.codigo
left join servicos s on o.cod_servico = s.codigo_servico

where pdo.codigo=120
and pdo.data_baixa='30-12-1899' """

    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp)
    # Resultado da consulta
    results = cursor.fetchall()

# Em caso de erro obtem e exibe o erro
except Exception as e:
    print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
dados = {
         'Descrição': [],
         'OS': [],
         'Placa': [],
         'Nome Cliente': [],
         'honorarios': [],
         'Validade': [],
         'Taxa Detran': [],
         'Reembolso': [],
         'Obrservação': [],
         'IPVA total': [],
         'Primeira parcela': [],
         'Segunda parcela': [],
         'Terceira parcela': [],
         'Total Dpvat': [],
         'Desconto': [],
         'Placas': [],
         'Lacre': [],
         'Outros': [],
         'Srpr': [],
         'Lacração': [],
         'Placas especial': [],
         'Licenciamento': [],
         'Valor correio': [],
         'Multas': [],
         'Valor B.O': [],
         'Total': [],
         'Taxa vostoria movel': [],
         'Solicitante': [],
         'Codigo serviço': [],
         'Codigo CP': [],
         'IPVA CP': [],
         'Licenciamento CP': [],
         'Multas CP': [],
         Re
        }
# Crie um DataFrame com os dados
df = pd.DataFrame(dados)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)

wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)

# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(caminho_completo)
    ws = wb.active
    for i, row in enumerate(results, start=primeira_linha_vazia):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(caminho_completo)
