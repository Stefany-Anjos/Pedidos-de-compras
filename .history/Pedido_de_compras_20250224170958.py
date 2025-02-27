import os
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
from openpyxl import load_workbook

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()
    
    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select o.placa, o.renavan, s.descricao, c.nome_cliente
        from os o
        left join manutencao m on (m.OS = o.numero_os)
        left join servicos s on s.codigo_servico = o.cod_servico
        left join cliente c on(c.codigo_cliente = o.cod_cliente)
        where m.data_envio <> '30/12/1899'
        and m.data_retorno = '30/12/1899'
        and o.estado <>'C'
        and s.setor = 'S'
        and s.codigo_servico in (47,65)
        order by s.codigo_servico asc """
        
    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp) 
    
    results = cursor.fetchall()
    
# Em caso de erro obtem e exibe o erro  
except Exception as e:
     print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha         
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
df = pd.read_excel(nome_planilha)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)
wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)



# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
         # Se não achar linha vazia pula o cabeçalho
         primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(nome_planilha)
    ws = wb.active
    for i, row in enumerate(results, start=primeira_linha_vazia):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
    wb.save(caminho_completo)



      