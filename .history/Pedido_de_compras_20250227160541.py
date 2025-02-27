import os
import re
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from email.mime.image import MIMEImage
import smtplib
import base64

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# Função para remover "CP" e os números após "CP"
def remover_cp(valor):
    if isinstance(valor, str):  # Verificar se o valor é uma string
        # Usar expressão regular para remover "CP" seguido de números
        return re.sub(r"CP\d+", "", valor).strip()  # Remove "CP" e os números
    return valor  # Se não for uma string, retorna o valor original

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()

    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select c.nome_cliente, (s.descricao)serviço, m.os, o.placa, orc.honorarios, orc.taxa_detran, (orc.vistoria)validacao, orc.vlr_correio, orc.ipva_total, orc.prim_parcela, orc.seg_parcela, orc.terc_parcela, orc.total_dpvat, orc.desconto, orc.placas, orc.lacre, orc.outros, orc.srpr, orc.lacracao,
orc.placas_esp, orc.licenciamento, orc.multas, orc.vlr_bo, orc.total, orc.txvistoriamovel, (cp.ipva)ipva_cp, (cp.licenciamento)lic_cp, (cp.multas)multas_cp, codigo_cp, (orc.extras)Reembolso, m.observacao, m.solicitante
from os o
left join cliente c on c.codigo_cliente = o.cod_cliente
left join orcamento orc on orc.os = numero_os
left join cp_os on orc.os = num_os
left join contas_pagar cp on cod_cp = codigo_cp
left join manutencao m on m.os = numero_os
left join pendencias_os pdo on m.os = pdo.os
left join pendencias pd on pdo.codigo = pd.codigo
left join servicos s on o.cod_servico = s.codigo_servico

where pdo.codigo=120
and pdo.data_baixa='30-12-1899' """

    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp)
    # Resultado da consulta
    results = cursor.fetchall()

# Em caso de erro obtem e exibe o erro
except Exception as e:
    print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
# Define o caminho completo da planilha
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
# Define o cabeçalho da planilha
dados = {
         'Nome cliente': [],
         'Serviço': [],
         'OS': [],
         'Placa': [],
         'Honorarios': [],
         'Taxa Detran': [],
         'Validação': [],
         'Valor correio': [],
         'IPVA total': [],
         'Primeira parcela': [],
         'Segunda parcela': [],
         'Terceira parcela': [],
         'Total Dpvat': [],
         'Desconto': [],
         'Placas': [],
         'Lacre': [],
         'Outros': [],
         'Srpr': [],
         'Lacração': [],
         'Placas especial': [],
         'Licenciamento': [],
         'Multas': [],
         'Valor BO': [],
         'Total': [],
         'Taxa vostoria movel': [],
         'IPVA CP': [],
         'Licenciamento CP': [],
         'Multas CP': [],
         'Código CP': [],
         'Reembolso': [],
#         'Descrição': [],
         'Obrservação': [],
         'Solicitante': []
        }

# Crie um DataFrame com os dados
df = pd.DataFrame(dados)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)

wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)

# Atribuir os valores ao cabeçalho
for col, valor in enumerate(dados, start=1):
    ws.cell(row=1, column=col, value=valor)

# Adicionar cor de fundo e estilo ao cabeçalho (linha 1)
cor_fundo = PatternFill(start_color="C3ECBC", end_color="C3ECBC", fill_type="solid")  # Amarelo
font_estilo = Font(bold=True, color="000000")  # Texto em negrito e preto

# Aplicar as mudanças nas células do cabeçalho (linha 1)
for col in range(1, len(dados) + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = cor_fundo  # Cor de fundo
    celula.font = font_estilo  # Estilo da fonte (negrito e preto)
# Salva a planilha após as mudanças    
wb.save(caminho_completo)
# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(caminho_completo)
    ws = wb.active
    # Percorres todas as linhas e acha a linha vazia
    for i, row in enumerate(results, start=primeira_linha_vazia):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
wb.save(caminho_completo)

# Iterar sobre as colunas da planilha (começando da última coluna para evitar problemas ao excluir)
for col in range(ws.max_column, 0, -1):  # Começa da última coluna e vai até a primeira
    # Verificar se todos os valores da coluna (exceto o cabeçalho) são 0 ou vazios
    all_zero_or_empty = True
    for row in range(2, ws.max_row + 1):  # Começa da linha 2 para ignorar o cabeçalho
        cell_value = ws.cell(row=row, column=col).value
        if cell_value not in (0, None):  # Verifica se o valor não é 0 nem vazio (None)
            all_zero_or_empty = False
            break
    # Exclui a coluna se todos os valores forem 0 ou vazios
    if all_zero_or_empty:
        ws.delete_cols(col)
# Salva as alterações feitas na planilha     
wb.save(caminho_completo)


#                                     ENVIAR EMAIL
imagem_path = r'M:\TI\IMAGENS\Imagem_email.png'
# Abrir o arquivo de imagem em modo binário
with open(imagem_path, 'rb') as image_file:
    # Converter a imagem para base64
    imagem_base64 = base64.b64encode(image_file.read()).decode('utf-8')

de = "robo@dukar.com.br"
para = "informatica7@dukar.com.br; localiza@dukar.com.br"
assunto = "Pedido de compras"
corpo = f""" <p>Prezados,</p><p>Segue em anexo a planilha de pedido de compras.</p><p><br></p><p>Atenciosamente,</p><table width="501" style="width: 375.65pt; border-collapse: collapse; mso-yfti-tbllook: 1184; mso-padding-alt: 0cm 0cm 0cm 0cm;" border="0" cellspacing="0" cellpadding="0"><tbody><tr style="height: 4pt; mso-yfti-irow: 0; mso-yfti-firstrow: yes; mso-yfti-lastrow: yes;"><td width="188" style="padding: 0cm 5.4pt; border: rgb(0, 0, 0); border-image: none; width: 140.85pt; height: 4pt; background-color: transparent;"><p style="margin: 0cm; mso-line-height-alt: 4.0pt;"><span style="color: rgb(0, 112, 192); mso-fareast-language: PT-BR;"><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAKoAAABKCAYAAAArMs8AAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAACluSURBVHhe7X0HfBzVtbfh5ZHwUVxkS7as4ootucjquyrbey+zvRftaqVV79JKq2JsUxMSEkIaCTy+fLz3kjzI98gjIYWaAgkhMZhiiqkhAUw1EJfzzhnPCsmWbEMgie39/37zm52555bZ+7/nnnPvnTsLsvjnhuaqsNhwc7dNeHOPs/6mXnPtl1rldX3hi7jgLLL4x8M9EbPxhsywfkQDK9J6yN9uhcKUCeqm/HtUbcwyTiyLLP6xcHQEVqm3R2F5mxSWjetg0bgKCvBc0CeHLcP6n23tkV/AiWaRxT8W8g63QHBpBHIHlVC4ywQXdvGhYEwOqya0sHFA/28ocs5RySyy+JTB72JSjTsCt1VNeX5YPeb8Dq+LWckFsZC0O80V43ZY0tMIhWkZLB8WQt6QCDaO6KE8ob2GE/tEUJq2nYf5L2nsNK+obdfn1YUNWXs4iwULbDbbv/CG7P+1elgHq7abkIAqqJ70/UYYcy3lRFhUtKpim0e0sBq1ae6wAArSCsjrkcAlnVrgd1hSnNhHhtzrvUDU462t73EONY76fi66PPa28KoYVI272KN2yvs+b9L/x6oBx2XlMUMVFy2LsxGSIWdexaT7gcVIvOVjOigc1EPNoO82XSz2fzgRFlUdyr6iPgnkp5Gso3JYnlJBcdoMW1M2KI9r4pzYKaEqZFnTkPT2G6aSD0vHY68pdrTuFVwae1Dy+ZYHanYFnirf4YU1o2Y0M4ywfEQHywZUULkrAJW91n+rChoLuWSyONtQ2W4rqt7pe6E4ZYS8ATWsHzbDtk7z17jgaWwZNO5c1CWAvDE1LBlVwqIRJUumbUTWZrWVE5sXUrdljb45OKHtTdzIT7haqmMMv67VmV8Z01GjYO1d6v5LejTFpQPGttLtrjdyhpTsiMOifgVcMmIGwbjnleqAfBvJZnEWQjjoXVcz7nl/zYgRlvZLYQ129ZtbpNu54Gms7dF/cyXap4uHFHBhSgZ52/WQMyCHqp0+qIiqeJzYcRDGPSJVMmiTR72ryeTgbp8QW5Py1TVT3j1Le9E2Htezw2S5PTKo3ul6a0tCuoYTy+JsQ10vw6++wnd4eVoJK8ZksC1tBLRPO7lgFjVtnovX9Rr+p3jCBEtR7iJ0sPJ2GmFhjxi2jtoPlLVo5tR2lbHYv3I/PxJKwuLiiku9B3KHVZAzpoKLRySwdATLNmX/FY9hzufEsjjbsHXYot9wpQUWDddB/oAIeBN2ULS5PFwwi+KAcPnmtPX3K4blsHhUCkvQucqf1EN+Sg9VE/5XefHZIwd/Kza3aewru2SwAs2NFZdqYdGYGFamlLCty9zCiWRx1gG75U1XMb9aPomaclgK67pVoO4N/A8XOo3CNulQEZJlORJ12RgSdkgGuUMa2JxywFavVMWJfSIQpoWfKR80P1QwKIe8cTlcONoIK8bRwUq7/lwWMC7ixLI4m7CpWzNSPKqGxYNiKEjpoGbK97Yk4dzKBbNY1SQSbr3SDTn9QiieVEP+qAJyBySwYcwMZZ36Wz9uN38ibEtqIusGlbC4tx7yd6EzNyKCNUMGqOg02TmRLM4W1LbrY2XDRljWJYG141aonPId3tikEnLBLGpCirLyXZ43cpGc+ZOoUVMiWIEk3Timh9JuzS/LAsJPRcOV+mRF5VimlSkFLBkTwdJxGaxEh299h/a/OJEszgZUoB1YljLB6iE1FPXrYOuIC3gJu54LZlEe1hRXjFv35Q1KYVFKAjlpKeQhUVfidfmo8Y8VTcoVJMeLKNbJ44xZ2GK7kI2IaOzy10g6vCbucoGwzVzQ2GMN1HQY+mTDPuWCdPpcLmhebBu1/roQHarFo0LWTl2e1kD5dvcHpzqKkMVpjsp2o6ZmygX5PVIoHtTAtpQTKtqYWQ5UZUy4dPOI4Q/F41rWJl06roBFQyJYndZC7aTj5a2t6ks40QW8pGlXTZcNKluME3S9Dkkp2hE/Ut3vBCIkr4tZp5wM768dswN/uweqhm2gSIduRNETrh1Y3am6fOmAGPKmlCxRF6dkUNSngk0RWQknksWZCl6SqanbEYY1g2Yo7NVAab8FattM3VwwCxqQLxsy/XRZvxQuGkaibNcgWdEuHVSD4IrmQ7UBRSUnymJNs/KKYtLK404o3+Fdl9+vvnrjhBXW9KhAmA58rjpluXRbrxa0V8YmnN8YrjJd3/eQ6rIENCaZ1VwSc6IoKfeykw1jMlg4LoGlE2h69EqhukU1q1FlcYaBukz9VOKhkk4LlPRaoXbMD5Xt5su5YBaqNtVn17bKbliJpKTZqPzLTXBhTz3k9Ulh26QXNkbUck50Gpv7LFfQ2tW1eJRMML/cuIOBgmEFrB1RA+8q5nz+LucNJX1KsH+rr8J6fXcd8+2UX3F5S+LYadtjUdyqFBWgLcyOMqQlsHhCgfYxErVNO86JZHGmosatEIvb3D5e3OzhRQ3mY+099OKvWTtmgZwxDeRMaWHpqBxW9oiBP+HCLldp4cRmoaRVe/WGlBUaLo9+r7TfCLWXB6HxS02PbBwxskQtn7LfUDZmBvO3his83x1/vWTEBtVpPyhT0U1cEnOipFlVvn7MxDpUy0YlsCQtZxsAv9vwDU4ki7MRtR3MpZsGGVg6rIULhpXsAur8fjlUjZhhS1AS5cSOQ12/+wuXtGHXviteo9vV/ELNANPHH3PfsqFbi0TtOr88zXxjy4gV1F/plNhvTjk1X+8/wB8LgLoruIVLYk6UopNWkrZA4YgCnTjUqGgCFKVU0Nhv/T4nksXZhnq/Tl3d5YDiLuxqBwxQMGWHYuzKN/ToYEtIMsKJzQlJt+/6hl4P2NKxooznrx4L3dY46GZt1C0d+uDWXhM4bxgF+ZUtd9dPhkG3q+2QdtC9mE1gHmwKKgpLx9GWRqLmjklZohaPaqFxwHY7J5LF2QZ5yFlr2N52QHp5Emoua/5rzZXNh6pGPR/Udpgv40TmhbTFlXBv73qaiTBLuFsLmIGmMftU6zNkWhBZxQPelHIy/ob5qi4w7Ur+Sthmn3dBSwY0718yYWYnGJYhURdi11+U0kBDD5MdSz2bwWMUS/jN5txSPG/0SXN4M4h3IqTT6XNtLS3TY6iEtra2zzJdsxeRkPPUkGC16Cm90rIqLN6wZsrEjt8umpTBonEFFAxpoL7LfAMnkkUW/3gURRsqC7cb4eJxCSyc4og6qAZ+p3EHJ3Ja49yqqqpCPl9cXFsrXCUWi4vr6+uLKivrixobGwt5PN7KhoaGxUKh8DOc/EmBcZZQevX1MjYNlUp1MRc0J1D+fMpXJpMVYT4FUqk0hwuaRnV19XI+n19cUyMsqKhoXHGy8lCadXV1+TU1NQVbtmwpyDwLF3yKSJ9biWUSizVs2ZTKozNL/6zIDTdIV0xq4IK0CC6eRBt1QgkFfUoojymbOZHTF42Nwl0aje55vd74ilKpflml0ryg1eqfo0Ot1j6P955VKlWPCgTCu0QiydewwixEXC76caitrV0vFIoflsuVe8Vi2VOYzvNiseTexsb5K7mhofFWLAfmo35Wo9G/IBCI94hEIikXvEAgEOi1Wt1jmM7TDQ2CJzHtp6qqaq/jgo8Dlk/I59f9Ahvc43K5/AmJRPI4pvc4kvvhioqKWk7shKisrPzXyvrGm+pkyqfrRdJn1Gr9i0YD87xAIJniRE4Ki90t0JmZL6n1pq9qjNbrVRbLdVKt/jrG5blOozN+zen0fcNmdX7darZfb7c6r/J4gt0eh18RYkIfa1+AohaZb9GQmB3sXzwlh5wxBTsztdErOG4s97RCaantPIVCeS8SEpAkYDSaAUkGBoMJ6B6SF3Q6A8hkCjCbrawMkvqgRCK9Cyuf4ZKZBSSJFOMcpniUHsXBdN7j84WbOZFZIEJgXu+QLOVJZ5PJAnV1DW2cyAI+vz5FZVEoVGyZ6EAu/pa0JicyDSRng1yuQmJR+fWARMXnoWeQHUbt+j2FQnFKdiQ2BItQIv9AptSBxWwHjUoPWrUOdBr9a9hwTmlKUqs3TzFO76NeX+gxXyCym3H6fu/1hR8yG20Ph3zRP5gtjj+YLI6HPd7QHwLBpkdsLt9eOjvMzh/4HD4L2bJcUqeE1W2KqfwJDeSMCiEnJWYXUK8fNECZX/KJrn/9u4NIgtryXiIVkYMOi4XBylWyZ9SuLCmIqAxjB9SOLPFsNgfeU76DxElzSU0Du0cl3v+A0iCyUNoY54BYrCjjRGbhaBnUb0ulcjZ9bAhs/kqlpp0TWYD5j1A6RFRqRFQu1KwPYtxZMzVlZWXb8HkeIrmj+SqR+EZMUwPY9d+OjeuUK6y+XvADk5EBvcYMBjWmIdeCWWdiySoQiGZNoc4Hs9Wxw+rwPmKzun5ms7q7jEam0+HwdJi15k6Pw9ftDkR6GF+gz+ML99kcnkG8vpRx+/4zGk484Xb5f2+y2pq4pE4J67qUtxNRF/fXQfGUEgrHtLBpyHJwQXrBRyL8Px2IJEi4+zKEIrIQEThyTWsvInCm8kmWyET38d5BHq9ulv3D5zcqMez9jIa0Wm1IPMUBtBfnfgUDy4DEfJPSI6JSPMoLu+ppovJ49UOY52EKp/LhcQTNhV/OJCqSdBXmcw/JUDlJjghK2hRt3gdQtogTPSnQZt+glKle0ymx4Ug1wBhtoFcZQCyUsD0Llu2u0tLS8zjxeWG3e3c6nL49flfwWr87uNPp8P9CrTb+xGVx36FRGu6Qag13GN2eH6u15jvNVudPrEhST6Rp1OEJjITD8b1uf/BRl99fwyV3QpTabOdtm7Q/l5OSsSTN7Ts6jVs97voNJ3L6grp+tDvvpz+fKniGBnxFoVD/HHvROxobBXdWV/MeRCK9QQQlGdKwGUIbDObHq6uFy7kkF6BTpJBIZO8R0UmeyI/y76JDVc6JzAI5Ragh36C0uLyJDIexa5/u+jGNIbFY+gHlTdoW5Q+Xl1fdi3E/R+FIwqVo495K8ek5yHyhZ1Kr1aj9BLvRETvh1OOxQPOlX6fRAqNHjaqgLt/A5mvFXkWpVpGWfg81dAUnPi9MFucujy/0pM8Z+IKLcX/b5Q487nB4H1RLNXdaTPY7DU7XnfVK1U9NZtt9LtSgVpf3UYvb+7gv1vzVppbkVbF48hW11nBKQ0tlDuGqDeMWWJiSwsq0FIpTcvT4lbC5V/8FTuT0BWkFvd50X6ZLzWhLJMUNWPkL0b67ADXhRUiI5XjWoJNzD2lX0n4ZDYhE/ytqvOlV5GgjYjTle5Rext7ENN9GLTUvUdEG3U/EonJQukjsQ3gkORHUcDVDmNZ7ZH5Qt4/lRHuzltVqGzZsuKimhncTkvQgxc+YL0R4tEcfw3zruGROCdigPosPcL9SrgCVVM1qUyk2TK0ZNb0RtapSDmotmhL1/Eu5KPPC5PBc5glFn/F5Ql9Cgt5k9/if0Ogt/yGTadezh8WyXmGxbDQYrFU6uS5usTjucfnDT3mjTU81tbR3hULx3bGmxJNOpzOPS3JeVIcUuhUj2O3v1EDumBhWDIpg1ZAaVifkNk7k9AVVNNqCrDNFpKODCFZby7uWE5kFJGFZppsmwpAGozMSexcnskAkkknRQTtA94ksJIMa7k2BQDovUVHudWogJEtnPA5inGmNio7QMGrMA1Q+CiciV1RU3rVt27ZldXX116DGfZfyyjQgIiyaAc/y+Xwll8QpA8vTgLbsO0a9CUxatE3JiTKaDqvR1pUgQbVmI0tUiVj8BJWdizYnLB7fFS5f+LloNPEVnz/yXac/9LRCa6J1pXPCpLcGfMHok6GmxCvxZMcXm+OtP/F7w8/ZbK4GTmReVLRqr6ZNLz6HJF02IYV89P43jTGw1itcx4mcvjhKVNW9VPlEggxR0FGZk6iIc7By7iNZIndGXigUfZMLZ4kqESsOaLC7zIQjsZCogjmJSlOGSLTXZ+aPJP/rTKKi8zJMZCQZagAYdqi2tu53qGm/iHm/dtQeZePRcxxB2T/V1NQZuegfCfh8l1E52OfTsI33iN5gurm+QfioAXsIDVcGzO+QoE5g4KLNCZvLe2UgHHshHE18Fb35Wzz+0LMmi+07XPBxMBic+eFI8+OBcNNfWju6boxEYt+PRJpesrh8Gk5kPpxTPe54gt5CvWBExL6KUjikAN6wew/ZrpzM6QuOqPdRJc8kChL1S5zIcUBH4o6MfZqRR7JM//lHiap8ZyZRUQu/UVcnnNOZmo+oqKVnERXJfiCTJ2lPDN8vFsteRe15iMqD4dQbHEFTZH9FRbWbi/qRQOYOj8f7PZksmXww/XckErkay7hTjdpVIdcAjQaolDoye27hos4JB0fUSFPL9cFQ7N99JyGq2WzfbLU697a0df6ls7P7mqam5p8nWltesno8Ik5kTpS0aIqLB1WwDO3SZRNKyJ9SQUGvBEqDqllrZ09bnICo82nUBWj33UHOSsZGJXkk0nR3Ng9R9/+tREUZlqiUb0aOiEQHXZONizJ4KF6kmTUu6kcC2rNqjUbz/sxnk8nku2mmDG30KiTqASIrkVSvM2Njkb0qk+nmzYslaij+fIaoXl/4GaOZ+TYXPAsMwywLBKNfiUabXw7HE3s7Orq725Mdz/qDgX1mj+eEM2Lrk4rk6kkj5KaVkDeuhEWDjVA6aoRNboWYEzm98XGIihr1xxmyzEVUrFRkqvJtreZDQhFR0ZOe9UpxBhmiklxGfh6ivpeRyeSfOWP6bDwqDzlaIpF43h7hREDb9JtmdJqI/NS9U/dfV9eQ+S/OaagTPkjaVKdFUwMbImlxdL4SXPhxQAfqSn849lw01vpVdIzmJGogEPicmWEMbo/v3/2B6IuJRPueRFvXeFtH5/fjseb9vkDgJk50XpSNMnsu6K5nZ6KWjsrYPVrL09Y/VepmjzOftpjPRq2vb/wyJ3Ic5iaqePrPJKJKJLJ3ZqZHRK2vF8854E8OCZLwNZLLyBNRRSLp9DgqdrHU9bNEzaRLZ3KaqJsmopKnT8Qi8mD+b1VXV09PwZ4KsBxL+fz6l+jZKB8aPTAaje8iedWcCA2TpYjApFGpIVKefH7DHfPNINntniv8odizGaJi1/+0ycxMDzeZTKYcTOPaWLzl0Ui06eVES8d3e3sHOwOR2K3RpvifopGmZ5qbm+e07TPgx3Qb16eMsGwcu37aJWVIBOvGNLCuSTivsjntcAKifoUTOQ5zd/2ziYrd799EVCTcwWOIOoJpvEfhGZISKfH8AWret4isGS1I5CGbVSAQ3jlzQuBkQHMhjGkfpueitCk9dAD3kN3KibCjHljWNzNaldPk++cb0bAyzst9waanp4nqCz9lMFmniYplTVmt9keaYokXhlNjO5ubk1+MRBN7XN7gvkg0vsdpc3o50XmxKSa7msZLac+pxWibLh4UQWlKB5vDAj4ncvrj4xAVCfTjjDP1IVGP6frRQ8+QisKxQt+Yr+vniMp2/Zn0iKioFTs4EZaoqCnfz6RJRMI0329sFL4ok8kGUXYvxSUtmLFblUrN+zPNh5MBe4qfZeISWYn8YrF41mA5/V/oON5N3T7ZqhnCYjnm3KzXYrHv8gQie6Po9XNd/16j8cOuH0n6/9Grf6o12XbP2PjkFeFI/BWfL/RiMBq/02i0nHSbSnmP/IK6nb736e3X3CkNXDgmZfef4o3b96FNdea8yz8fUdFGnW9l0rlYUfeTLKfRSP4IEvVbXDiNeUqpm55JVLx+a76ZHBpgx/D9mTLQGTXjoZlERQ2eQjK/nwlHMh3B8P3YVXcQ0Wtr+SGMc4A03MxGhER9ArXiei6ZeUHTu0hUduaN4pNJQUTHtI/ztnm8umGzycYSlUwA9lBpH5xrgQx68Ts9vsgT4Ujz9cFw7BaPL/ikwWiZJqrFYv0Pny/wZLKt/YGBoZHPNyfa34w2texz+/2n5ASVxORe2hCjIK2Gi1CT0rK+whE1XBIWnFmbo6Ejc55EIr2XKpW6TKog+o0OxFc5kVmgsVAk1ZskQxVKWoe6W/R+r+JE0CFRNioUatZzzhAV5T9AW256F5CZQBuwFEn3NmkySpPSxjwOIemmd2lGwqUw/AMKowMbyRHk0MNoh7LrVo+uPa3/IYVROlQulnR6C1RW8ua1tzPAHmQCn/0w5U+HXKZGja3aQ42AE5kGef9KpXY/adXMgeV/Sy6XCziRaRhMzKV2t39PMBS/LhBq+n9ud+Dx2URlkna761G0UZ/r6RtINSfaHkq0tL9id/u+cLKVU1S2+knXH/MGaTt2GdAWQkuGZFA+5Tyy1XuGfZGFNKpGo70X/2i2gqhyiWBIku/Q1CkRgOw87LaX1dbWYl0ofpMhMxGUtCpquiNIVB+XJGnUErx3gAjDkZSVE4slP+Tz+bmcGAuaokWS3Egy1G1T2rSIBc/Y9cvNnBhLVMzvryRHB6WLnv0fampqphdky2SqBrVaxw5zIdFZGZlUhUQyojYXSzix40BlEAolvyf5zEFaErvzOadIqQdoaBDdTcNT1BDowIZxqKKi6hpOZBp6o3XK5vDu9gebrvX6wzc7nb7HdIYPXwshZ4phbHfGm1tejDcnHmht7eiMNyf/7As3PefyBpyc2JxobNZLtoyY2U2F88aUUDxhgFVDWtjYrj6uHKc9aFEKmpT3EUmJeKQBiajoDL2iUCh/gU7ST5CEP0Mb8SEk71sUPlPzUqViJT1Nq+i5JImoC9FbfowWhFBaJMtWvkpzuLqa9zNMrxe7ch9q7S4kw38joQ9Supn8keT4W/0ntA83cEmyREWyTxOVDmxcuykvToRwbkOD8Atms/UQ5cc+k1oPeoMF6hsFv9q6deucWobXwFPJFap39SivxDLQ7JNCqT0wn4NEwDIO4XGQHUsVyVktjtd7dTp2C/NpaPXmCYvN87DHF77G7Q3eZHd49yBRZw1P2e12gcvt2Y3O00ttHd3Xt7R0XNeUaH0t1JR4IBKJzLvuVThs/2lhjwzW7TTDwn4hLO0W0UopKIsqP9ICnNMCqC3/Vas13E+E4jQpe86QkMhD16ShMt04HSRHA+wodwhJfNyHFNDeG0GisounKT5pS0qTq9CDSDyyYY9k7hGpMhqV1rri+YaZXR/GGZ1JVEpXJpM/cgxRqVvegPYyOixHh6yIpHKFBkxm6+G6xsY+TmwWxFLpdSrU+ERULaZLZJXKVL+eq9vPoKKitpIaMzlSpFGp8WIZDzQ2ihycCAut1jhuYly/c7n8n8fjRpvd/eixRCW4XJ5ul8f3LJL12c7O3t5oc/Ie1Kp/Zuzum2Jz7I7CS5pqtg6Z2I+vLRuUQPGUHtYMa6CyQ39mvr9PRMXu/H4iHxGKzlTJRBiWXBojWxEURveInERajnwf4O+xuWwprOSl6P3fTgQksyJDLkozQ166R2eSIdMgU4aGhsaH0PGatccS5YMyB0nmQ6Iq9xxLVAI6gu0oc5gtvw6JZLQi+XQglij3oeYv5cRYYD4rZHLlU0RUIqlOb2ZlJTLVCd/Lp2cWCIS/IBOBnCqlAuNrjIexAd7GibBQq/XjRovjt3an73qHy/9tm92ze6aNmgH6Cgt9geAt6Ej9pSnW8vBAaqwj1tr2XCze9mo4GBvgxKYhGPfeRVOmRWkN5KF9mtcnhi3DJihzi+o5kTML5EyhjfpLIgCRhSqXiEjL6YiYjNUJEvFRbcqZB0eQrAfJy8Vuds5XUTJAEhWh7fc9jPdX0ppEMDoTGakxkLama8qXI+sh1JI/UigUx632IaJmykgHpYX3HpuLqGRXozN3Pz0HduFs98+dD6MnP714hiCUSJyoQQ9Rd0+alDQwyqJNKznpWlM0XboVcu375EyxY6oy9WFsJK9gmaZHGVQaw6TOZHsANePlNpv7W4zd9chcRCWg/VoebUo8FgzFXmtKJL/f0Tvw5baO3jfCkeaXvN7g9DtP6mZGsmXwqG1Kr5qsmtJBYb8CqgZNP+ZEzkyIRNIrkKzPaLW6Z7CCn0HyPIXnvXQIBbK9GrX+MayE3/FrG3+CDs5X6uoaHeiAzHKK5gPZbJiOH+3g241G08tItveQaETcg0jOg6SVkbB/wQq+AwkaQ6dtzpcGMcyEck9gvGfxTO9D7UON+k2afuVEZoFGJ7Ax3aXVm/YpVdp9ZovtGZVa9zTa3D8g54kTo2dvV6rUT8iV6ufQVHhWqdI/JZYovkwOEycyL7DXKBAKpHcKGsWP4PkRuUzzGGrUB5Go00vylBpjXGdg7rZYnGmL1fFFi8Xxa7Rbj/uSSgYujz+aaO14ORBueql/aHRXa3vP3a3Jzrc7Wjp/3Z1IsN+PMvaFrto0YGK/Y0Vjpwv7BbCJ5vUjZ6g2zYDIhJW3mjRZ5sBKWEeagRb31tQ0rpZItCvn0l6nCoorkUgqkLQOdKS60J4bIqcKie+VyzW1mN8Jd2EmQmI5i6hsSKK1CoWeynfC8qDpkYPyG+mQStWXKJXKDRKJftYCZHp2lFvDLmLWatfL5frVM4l8MlAeYrFqrVCoWCcQyNbjc6wic4oLxue2LdQYrLUWi2WjXm/dajRa6zQazfTbEMeCejin2/f1eEv7n0OxxO727oGOZFvXc+3x9ncj/uB1trTtPH27M09yWezt/AEN5A4roGBYBRVDplu5JLLI4u8DhyNY6PSFfole/8uRRNut7e29oy2x1te6k52v+ZoiAZLhtzuk24ZdsLZHB2VDDNQk9SfcNzWLLD4V+HxRZTDa/JQvGH2pta1rZ0ei47uJePINfzT6qKO1ld25Wtjnv6yq3Qb8Tvu8psQZiWO895PudUTrJ3U6k1qv/7A7ncdm/Ht/Cvwj5XeyGSAOM9M8Nv1ZYSdJ71TLdo7N6hqNNjU/5/OHHku2tCdjsdYHE8mOvxjcrltoTyptwr1Y2xW8pSHEfKzNKk47kN3mdDq1aEex88sajUFgtTrUDoejBIl3oclkNSoUWvaVXYPBIEPPOYiya8gWMxjMrVqtlnWAjEaLEuMp0IacHig3m80rjDqj24SEplkYSg/vEbROvTOPTV9vshg1GvYFPKfVyu5iYjcf3bkOwxK0eRg7g2NhInRPrda5Mb4c4y63272bqaxoG35GpTKUGxhmg83mbMYyse9K6fVmOcqyaZnt9mo6z4TdbtebrFaFymZiHSCjy6VXm800izVNKF8o1GBmGK3H4yllIpEljMvl1+j1Rq3bvZgaq8nhMGr0JiNLHpOp0mi3t6AM+38xjMNm1pnZEQSdw6qW4v+H5TklR9Tr9ea63b4fBEPhfW63/6eJRNtQU0tyX7St41VHJDJKMrFP4TNA/7SwWq1S/PMK6E+na/xz3UjccvyjLsB7K5GAPpstwDoA5Hxg5bKresgxsFqdVUQGJHA+nem+Vqtn7SgCORE2k6nBYTKtdVgdUhcSg9L0YH4uu0uJjUGBxypOfIHVZGWXtTFmM7vnvMVkSSMhG9wOt87BONhNH5CIXR5PmJ21wrhY96ZuhgktQyKxU65mM+OhshN5bVanHeVddF9pYr+3NEuj2Ry2IJ0Zr5dtBCjT7HK5Zi1iwYZqj0ajl9ACZ6PDUWj3ek3OYFCC5zJsGIwxEFjkCQYrjG43+waDxmhk08LyXIzPaGIYO/ssSqupWelhqNynqlWpbmqj8dijkXjildbOnu/Ek13fTnb3729ubn8+2tz6kV9cPK2BFYP1Y8rBLjyJv5fq9Sav3e6uNxjCF1E4kmGt0WiMyuXeC5A056Ec+9KcTmeMqlQ6JIq5lWG8q5G4jXRfrdb66UzAdEuQ2DxMo9BitGiQiK5wOHwRaQv8bURCmme+Cmxn7CH2bD969jg9YY/HF3DanGEH42YJgNwYwWikpc7B3/JIJFJDjQsJyr4WbHM67aRpqPHhc/QzjDONDekiq8dDjWAWSewuF5uPye1mdyRxhkJdtkBg1iszdrsr6veHGCTqckcgsMrp908anc7t1BiQuNiIbf/i8PlKzB4Pq7kxnzCdqeEiUZFrNlb7YVmXWPx+gRp7ALo+VRjc1tZgsvnPoWTLi92DqZHmlq5HOjv6324KJu5t8jT9U2/a9okCteRGt9utwz+0Gc/YnTmxV/OLsftcT+OJDOPio1Zx4R99PmkprAB2gQeSF3s9RyHGR63BnI8mgQuJ2SCTqabfmCQTAdPxI7ntKLOOtKfFaPThYXIzzBafz1ek02h8SMxqqnDGzDiNWqPIZrWhJsRrxmm2W+xKDBfgmV19hQRsRW3ViATPt9s9PLxegddhNAHYFf2M0ymhuGazvczp9G+1WOwC1Iir9dhFm1wuIYWRHMHIMD6Xzye0IrnpGrVkyOaa/Woykl1Gz0e/iawoW+uPxQRI6OWeUKiU8fvNjNfloWtWHs2Ho7IRFf6fxZ6AR0Rlxby2ODwBkRnLTOEng7HNUWdKumpUaPpY4/4bfe3N+72JxJ6R9FR3S0v3Cy2J7jeDvujVp2hjnxlAsi3FCmdtJ9JGpIHYAAQ5SzOvuYo+JzMojtf0Ou45NP4ok8nyj3XKaNwQST09TkraO5MXAcmak4lDhPD7/SupQdA1d3+WFqT4mCeRYvr+XHLHgtKk7pu7ZEH3XLHYUu6SfX/po9p9ZLdieWZuyjtnOejZqMfiLk+IhlbTJdbLWw8Zd8Tflfd4c5XNzAZXR9MDka72t6LJjh919Q3vDDe1vh4MNb0QDEZO9ip1Fll88mhIaBc3pl27N/ZpYVWPGqrGnN8hrapKuIS+1vhtkUTiHk8wHPY1xa/1xuO/MDk92W+dnsU4oXb+tMDu3T/lv3v9gBqW98tgRVoPq1MmKO+yXErbppPGJxufejHq2TImSRZnKewd4TnfRvg0QV9HqRphflDQLYXcHgGsGFfDBWNyWDKshvoxD8iTznkXf2dxlsKdar1NeIw9+2miri980aZ+w/9d2SuDwnENFKQVkJOSwsVp1KwjOqgf8X79rHKasjg1aAcjz4g6nKe0hfrfirIO46KyYea/14xoIX9MxX4eMj8th9whCRSkDLBh0PLdM2L/qCw+eRi3J/dJB8Mn3Ffqk8CmYGNh7bjrt7RSf3nq6C4nS0clsGiwHopTSijpMfxnceDoHrBZZHEc1Ntbf8PrdoEgaDnpq9YfF7ykRVabdr1+SdoMK0ZUQB+JyMUun75hWjipgOIe6U085vhXsLPIYhqVHZZvbeuygqjH96PKT3g+XRgwLqpO6kfLU1YoGlDC4h4RrERbNH9UCatSatgwpoXCDuFlWU2axUlRGdcnSzoNsHHUCmu7lEPc7b8N6AxtalboKkbNu9endexnIPPSMsgZFMG6cT3QkFTdiO3w1qiEnSrOIouToiauLdvab4acPgls2GWFkk5FL97+WGOrNPtVEpMLqkaZn9J3Slek5HDhYAPkTMggF7v4ogkNFHaJoKxL+9hWr+jv4sBlcYaABtN5g86nV4/qYFm/CNZ2S6GiT38beuhz7vE6B87ZElUVrG0S+/hj7l/TRhEFvTK2e8+b1LBffV6Cx8KeRtiyww5rWyXf3BZSnR3rSrP4ZFHTboysG9ZCbq8Q1g6poARJWzpqhpJ+/W2l7drYJSFZQ2VSt3FrVL66PKJYVxqVb9sWV2k2NSv7KlK220tHmXdXD2phbUrHzjLl9IrRFtXCsiE5rJ4ysV993jJufbI4JrZgdv+QmbAszgCsa1N9tmzE8uA6cm76pFA0KGffoV8xLIeNY3rYOmoC3pQT+DvcUHOZB7btcMA6tDNp07JFXQJYntawG+tejN58Ab3SjHbpygEFrOpXQ+2U9+AlccVYZUx3SotUssjihNjSqt5Su8N9oLiP3vZUYJcth6Xb1bBkQAgrh2SwCu+tHJRC7oiU3ZeUxkKLLjXAhUMiyNupZ+WXT6A8evZF/QqoHnfAtnbttbVx7ac27JXFWYqypLpeenXs9fwBFeTs0MLnpmRw0agElo/LIT8lZY+VUypYMXl0//yFI2K4EL35hWk5FI/SCig5CK8IvVvTY75qk1+8lks2iyw+efB7LMXbhpmvlV7hObJ8HE2BcQ3koyYtHpRA4YCY9eSXj8hYrZs/KINNV7qg6grf4coB648a+22BzU7JST9olkUWnxiqOi1rantMMcGw/UbpuHu38srgq+LP+9+qvdL7Us0uz0NVI9abq7sNg7VxtZwX157eX3TO4swBzVqR01UZ+3B3lLMPCxb8L3m1Nyc7NtE0AAAAAElFTkSuQmCC="></span></p></td><td width="188" style="padding: 0cm 5.4pt; border: rgb(0, 0, 0); border-image: none; width: 141pt; height: 4pt; background-color: transparent;"><p style="margin: 0cm; line-height: 105%;"><b><span style="color: rgb(64, 64, 64); line-height: 105%; font-size: 8pt; mso-fareast-language: PT-BR;"><font face="Verdana">Robô</font> <font face="Verdana">Dukar</font><b><span style="color: rgb(64, 64, 64); line-height: 105%; font-size: 8pt; mso-fareast-language: PT-BR;"><br></span></b></span></b></p><font face="Times New Roman"><p style="margin: 0cm; line-height: 105%;"><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>(31) 2555-7850<span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'><br></span></span></p></font><p style="margin: 0cm; line-height: 105%;"><font face="Times New Roman">  <span style="mso-fareast-language: PT-BR;"><a href="mailto:informatica1@dukar.com.br"><span style="line-height: 105%; font-size: 8pt;"><font face="Verdana">informatica1@dukar.com.br</font></span></a></span></font></p><p style="margin: 0cm; line-height: 105%;"><font face="Times New Roman"><span style="mso-fareast-language: PT-BR;"><strong><font color="#404040" face="Verdana" size="1">Email gerado automaticamente.</font></strong></span></font></p><font face="Times New Roman"></font></td><td width="125" valign="top" style="padding: 0cm; border: rgb(0, 0, 0); border-image: none; width: 93.8pt; height: 4pt; background-color: transparent;"><font face="Times New Roman"></font><p style="margin: 0cm 0cm 0cm 1.7pt; line-height: 105%; text-indent: 7.1pt;"><b><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>&nbsp;</span></b></p><font face="Times New Roman"></font><p style="margin: 0cm 0cm 0cm 1.7pt; line-height: 105%; text-indent: 7.1pt;"><b><span style='color: rgb(64, 64, 64); line-height: 105%; font-family: "Verdana",sans-serif; font-size: 8pt; mso-fareast-language: PT-BR;'>&nbsp;&nbsp;&nbsp;<img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAADEAAAAxCAYAAABznEEcAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsQAAA7EAZUrDhsAAA6/SURBVGhDzVoJdFRVEk0URwUVHUUMSELS/X//LOASXCYqcQENWbv/kqT3JSEHEEacwVFxJDi4IIpHFFQYICFJ70kM+06M4MgIIzrDQQY4DqggyKLEDUax59br14EsIox4zD3nn9+vfr33qupV1av3f8eda6TN1C4R/MZUXcB0Y2q9nEQ0wSsPE/3yJqmxZLXoN04S6+TbGHN3g+iVJZ3f+JhUrzaLAXmfGFAOSUH1gOgztYhetVwIyLlSg9aSscoWSV9hjRiCSqO0wHgl7949kBbSEiVY3xBUDbr5rgEpaNOKQIm79T5llBDIHUp8hoCqSSGtGavyPZRdCVK8wVtUIIWV/LiKzAuIp1sCCvWGG42A9adDgXoouszgk9/E6rQKAZNDqoXyYe3z1MXmTaQQ79Z9kDm74gLBZzJB4PVSSP3aEFC+h5u1QpFdYkh5W/DL84gPcZObtsQcSV9miQi1psmsc3eEPqj2EWvUZLjeNQmz83tSW6iXU/Qzci5kz2vgVk0ln6YttUT0PtNwkOKJ3q2R7DeKWIm/SmF1uxhU6KrT1RlnpTYVfwTX8+tqCvSIqWkps4f15l26JwbWyUlCUJkMV3sH17dSvRZBwO/V++X1Sf6igQa/8rFYr23WVxt1vEv3BrkZ4iXb4NUmIqNNT1iU31P0m6YMXu+mlLwlpcZ0NWc9M+TnV/R0uVxZTqfTaLO5CqxW5502m+06h8ORbrG4UyorK8/jrOcUSd58iTbKpJm511A70Wu5AoodHLTOEYFrvcyYfgoQtJfD4XoU13Kn0/2iw+E22WyejNLS0n54lmW321OhVBaUMePqz7v9X7BYnNdjrAlkME5C8BemCV7T+4Z69SA2x9eR1Ybp6kwT0pdjgwzIu6gi4KxdA1Yf6HS6/m63Oxd4PJ4MTmbAZI9CqX9Yre40alut1mttNuefysrKfssYzhIWiycJhtpvt7snZ2ZWtNvkWCwE1Y2DWpws5Yo+eT6utxAfX+jn5fThbF0Dln0GA8/kzTZAuXGgf2W1uu7gJAbQH7fbXRbePGNomnY+VvhNGKyWkzqB0i7KlTHYHD+IKqLsQv21kQJc8BXewtk6A+5yi9nslHiTAXFxByY7CoHv4aQ2gL4Yyo3lzU7Aal5aWZndgzdjiIcCD6LfHk0bc3rXIETi4qVGZaToV3YgHc8SA6abkL12UUnDOU4PCH45BN2MCcs5qQ0I9DzQW91udx+Xa/QA/H4C118QM0NKS8v6QtBZDodzEVxOIX6s8r1w0/G47BjzEMbOIbrHY8uI9nNOQYxkEi2GzM3Y7YPGm+l3ijdPSPEabxe8xpvTFpspPhoYU1fAUv8mAu3pN+IAgjk6uRdoN2Pir3GfQBPj9zZcr4L/eauVhHX5cM2gjGax2PMh4GQ8XxPNbu4GtMPRcdwq4mwH2i9CqRcsFsfQ7OxTVg5yoBp+2dCoLkmcMzyZSEKdPDS1oQQpVz2R7FMGM74YyKLkSpjoalLCavUYIEiVqqoXcxYGp9Njdbncx2hiKNEfQhyAgDOhQAH4cyngsRpOPFM1zZ6IZy/B8jspq0Ho0eBHMNsTwYIVcZ/A80ooWoQVywFdbKcEQG6T2lAcSV1YclT0GYt1voJ0KaCcoNXQe01Wzsb8/iYMcJ+qnsw0aGuge3iT4iUDwix0u8tbMfl4opEAuHZbLJZBpEBpqX0I7kVwo1V2u2MNDJAMwT8HzzN2u4ee+fBsbXGx6ybQV4H+Nilns5UVYCWup4Bnk3WA6DVOGrTWEZEaio/rEezIUh+mo74SfEYzY7DZ3LfE/LYjKDBh4Szo4nO7yz7GPRxLsQQIMASCPAue4ZQyKVVivGooOQtCSfn5+T1h8UkQ/j74RjxoeWg/BqWvoCSC1XoSY2TTivIhu0Zzdg9kpjmkCO7vIbB3okQ5pptfNIA9t9kcyP/MCr/BdQmsoscKYBU8z8LyLbhvgQIvkSVZh18RBr+8AKX8dziLfIGT5COGuVmXDqwtNCDLuG4ka2KZkU1cc+HvVfhdDeGRMVxFUPByPsavBhyaEqQgK83jUHq8iVS7nP32m6r1AeVJ+k2Ip52XBO4YWN0Bhirs4AH5E31twa1CdV5/saYAZ5L8RFS8J0B/g7N1byR6864gN5Lq1c8Eb14K0agYjL5wUD9mTN0eIe18lOJvD2p2Rgw+ZTVTyidvowyFfWQ75/pRxFOG4r/bYcyYMdfY7WWpZrP5Kk5qB+wXacg8ibz5syH4lYnM8gHlqxRvYSmUOJyx0kbl+YucpTOo9EYuX4v0Z+MkBrQLEfhNCPqFSJd/hKDyqcFPKRjPlyJJzMDz8arq7tMx/7tcdg1pvZQyXjSFuzzIhBujqbhrCMG8/nCdAzhb0MlvDm2AKNe/MQSRnboCrHgZBn0P1yuUeolmNrsEj6dsBSbaSjsvhL/SbLbfjjLjgRgPJQUIvx7CT6U2AWXJbVRL8SYZYTDG3Yz++UjvDpTiT6PPq7gm0G7P2dohubqAlRf6gFyJQD6BYnB3ar3WmuIz3c0YOoIEgmGaMdF6TqJC8B6UG0ewW2MjO2l1bGpUTrfwJgR2p6C9n1yN2rAwDlHOkTElqYSBsBtiO/6ZQF+bcxlcpoWOrjpv/u+QXg9DiRZ9bWHaVa/dl6DzmQo5axQ0GawdxOZ2HDuqnmi8aPsWkz/BmDig1LgxY8ZG6M5JcTk5ORdCiX+Bt5yEp0NTTGmM3RvCN+H5N7D+MCp16HzSsTbrCMmnmDPW2COGWuMDfWtsvXD29lwbVC+WFhhFqbH4I5TmbftEHO0TtOGRYJgkZimUCm74tzPA2wxQSoWgH0LZI3CBBE5mCmMVd6LPUbhNMSczoD0CCq+lQpHcNboqVA07p5NrcrZOQJkxj5cbm+NCcSy2dD4l3RDW9mWsYfQSxogKNg1Wux/CNePaAWF6EZ3uaG+x28tFalN9BL4K+PJYCErHWB/RKQvBsjUQcqPTWYad373HZht1Rm8mYIxmjHmyGu0AZKPllEpRK/0gevMlfTCnDwJ8N532cI8keWWJBE3GIDlkRQh2BL9LeX8GtF0I5j8QnbIILoECErxfW60jEyD0bSDvg+CBWCxAsI24nmcDAHCl8+FWF/FmGygDwr2OYHyZkzoBmWhx+hIzlNB+EGuL7kE8LEpbYomevfGMMZEf0x010zQM+O/Kysq2siMrK4vtEVj2/qWlnn70m84ZUOC/UIQyVCL6fIkVeIiexUCHf/C8Ff1tvwFC5lJMsIcncR7mXAPlP7Jax13GaW0gl0mqyr6I0mnGKju5zT9Rev85takkQkoZwupRoa7oes4ejQdMehDuYeckEuT2kpKSdudYCJTqcpV/BsHZAR/WnksXe3gKINwc8KyE79+BMR/TtIp2CtAKIJ7WYAWPUVxwchtw2HHjzLCWdmrJrzxPr2tQ7K2Dax1mQR5SdwvewmGcPQqHw1OISXdmZUV354qKit44F4xmDzlgMQcmPoRrHrlHlOaooiMlY+CA2+RAsW/JReB01VCcxROBVg7x9zhS9X6MsxW87c7UBL23SKG34xB4FbVRVnihSCuUOCLVF283BOUHUl7p4k0gJh2LCVfzZiyfT4PgOE566HXNcrfbswX3di8MysvLb0Ws/J4OVvQbQj0Eng3YABkflHwSxgmj78u0MpS5kMI3ejzl4yids0FOgX5B4Q3w/VZyHymo1IMULwTkqTiiviDUmkwD5hb2ExeWFAkBUymdv6O9OGinpOA+9dUkLJaCSSfCzabAqlosdjoCCeFq9DViZx4JJRQ6tfFHcTQe6isNQj+MZ6Mw5nX8UZcQfQhiZKL0ZXCfQPS7BYF9EvDLo6Wm4q2prxfDtYo6vT76VYGdmK0IvZqhNJraWBxJW1gakXym4Wy3DuEEF1Y/JeXYfuE1TWEduwOues2UIPjkR2NWheWforSZtsgcAb0uGYGLQ8/7FB9Eo+BGyVEbVxn3i7zIPivoZ8C6AbUSwn0CoT40bMi6lL7pGULa7rTFpScMftMG0GdiR/6KuRb2BFodKDkjtmqdIAYLkqUGU0I7DZuze9DgfWuG94JjnzPNUxaMGILy+V3yeToLCCHZRXTUP1Wog44jA+0Rg8oBWhESHgeh74Sguoy+57EBfgx6v/FewWd6S4AF0OlvuN5BevsA9z2478b9PZS/TbDQs/BNjb0mOY1i2TAAfZ4SllhS6HMW5Xqi0+dhKaztpY2KBEQQ/4c2M32t6VoDvWP1m57GZjZVDKmVQlAeK/oUm1BjOm0iOIlI5XlQ4gks1xEIvd8QUL+jZcyAD5K1MlbbI3QspJ0SSrTofcZRtEq8N4MuZMwUg9rDUDaE/u9KDcVfQJhNMX8fWJebJIWV7TQuvbnDWeAgaNH9pTK7RyIO/un+3CyxzjgCijqhjJM9OxuQryGFVWPyVgi6VgwrftTwVbBQLX0MNIRLJgoh7eQWHwXl8KFSQPUa6pXj7PMtFKcvOjjQb6KPhsREL4Cxw7LX8/TpV6zXVqKs9hgCJg0Kz0ERtxHG+xzCR9JhNENY+RLnaA8Zl81yVsCyYzUa6RyLAQ9A+IU0ES0554jLnJ15gS4g3wWFHxED2mapQWMrlg7r0gUDfCMG5fkUvMSPtGnG6hyEYN9j5/0UxdsbmKMZQbyPViQDc7ELwmN1qJxoYi/Dfg7owwasMDd9KSbArpkGF4JC2yS/PLbvcwhy4vGZxrOsElIPQaC9UGi7FFJWwKpThZiLcLAvo9hdMeZ0JjzclSxO2QY7M5UWx0Dfg9WogbJ38W7nBkKd7GIpDy5CMUHZBFllK6zMDiD0ZxS4XJUUVo8g8A9LAfl93Fsg1HpYex2uMNLoU/S97dQ/oST7CvqSW+r8RXfqg8ZsqkL7/ZJ/UqFvY8gQ98Ot1sHqx8jfyWfpXzRJ0RI4HgkhDwG8LhU7LK0cuQUpzHhXRP9NQ5ZH0bYNFn/uFxX4p0BplawKN6qAUtP0fnlS/9qctjcTOi+yk095EG6B7KR8gOJtL9xrB9pLsQdMQx8bng9Oau58MPr5iIv7H+u+7vKhO9wlAAAAAElFTkSuQmCC"></span></b></p><font face="Times New Roman"></font></td></tr></tbody></table><p><b><span style='color: rgb(64, 64, 64); font-family: "Verdana",sans-serif; font-size: 3pt; mso-fareast-language: PT-BR;'><b><span style="color: rgb(64, 64, 64); font-size: 7.5pt; mso-fareast-language: PT-BR;"><font face="Times New Roman"><img width="495" height="3" style="width: 5.156in; height: 0.031in;" src="data:image/png;base64,R0lGODlh7wEDAHcAMSH+GlNvZnR3YXJlOiBNaWNyb3NvZnQgT2ZmaWNlACH5BAEAAAAALAAAAADtAQEAgQAAAKampqWlpQECAwIVhI6py+0Po5y02ouz3rz7D4bi+AEFADs=" border="0" v:shapes="Imagem_x0020_6"></font></span></b></span></b></p><p><b><span style="color: rgb(64, 64, 64); font-size: 3pt; mso-fareast-language: PT-BR;"><b><span style="color: rgb(64, 64, 64); font-size: 7.5pt; mso-fareast-language: PT-BR;"><font face="Verdana"><b><span style="font-size: 7.5pt; mso-fareast-language: PT-BR;">Rua Dr. Gordiano, 164 | Prado | Belo Horizonte / MG | 30411-080 | </span></b><span style="mso-fareast-language: PT-BR;"><a href="http://www.dukar.com.br"><b><span style="font-size: 7.5pt;"><font color"#0563c1"="">www.dukar.com.br</font></span></b></a></span></font></span></b></span></b></p> """

# Configuração  do servidor SMTP do Outlook
smtp_server = "mail.dukar.com.br"
smtp_port = 465
senha = "0za*s,}mx}e."

# Criação do objeto de mensagem
msg = MIMEMultipart()
msg['From'] = de
msg['To'] = para
msg['Subject'] = assunto

# Adicionando o corpo do e-mail
msg.attach(MIMEText(corpo, 'html'))

# Anexando a planilha de CPs para autorização
with open(caminho_completo, 'rb') as anexo:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(anexo.read())
    encoders.encode_base64(part)  # Codifica o arquivo em base64
    part.add_header(
        'Content-Disposition',
        f'attachment; filename="Pedido de compra.xlsx"'
    )
    msg.attach(part)

# Conexão com o servidor SMTP e envio do e-mail
try:
    # Conectar ao servidor SMTP
    server = smtplib.SMTP_SSL(smtp_server, smtp_port)
    # server.starttls()  # Inicia a criptografia
    server.login(de, senha)  # Faz o login no servidor SMTP

    # Enviar o e-mail
    server.sendmail(de, para.split(';'), msg.as_string())

    print("E-mail enviado com sucesso!")

    # Fechar a conexão
    server.quit()
    
    print("Robô finalizado")
except Exception as e:
    print(f"Ocorreu um erro ao enviar o e-mail: {e}")
        