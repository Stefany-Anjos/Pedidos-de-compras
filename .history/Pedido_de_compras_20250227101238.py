import os
import re
import psycopg2
from selenium.common.exceptions import TimeoutException
import datetime
from datetime import datetime
import openpyxl.workbook
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font

# Pega apenas a data atual (sem hora)
data_atual = datetime.now().date()
# Formata a data para o padrão dia, mês e ano
data_atual = f"{data_atual.strftime('%d-%m-%y')}"
print(f"Data atual: {data_atual}")

# Obtem o caminho da planilha
caminho_planilha = r"C:\Users\stefany\Desktop\Teste CP"

# Função para remover "CP" e os números após "CP"
def remover_cp(valor):
    if isinstance(valor, str):  # Verificar se o valor é uma string
        # Usar expressão regular para remover "CP" seguido de números
        return re.sub(r"CP\d+", "", valor).strip()  # Remove "CP" e os números
    return valor  # Se não for uma string, retorna o valor original

# conexaão com o banco de dados
dbname = 'dukarsys'
user = 'postgres'
senha = 'S@ntos67'
host = '192.168.20.5'
port = '5432'

try:

    # configurações para o banco de dados
    conn = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=senha,
        host=host,
        port=port,
        options="-c client_encoding=UTF8"
    )

    # Criando um cursor
    cursor = conn.cursor()

    # Configura a query para buscar os pedidos de compra do banco de dados
    query_pedidoComp = """ select c.nome_cliente, (s.descricao)serviço, m.os, o.placa, orc.honorarios, orc.taxa_detran, (orc.vistoria)validacao, orc.vlr_correio, orc.ipva_total, orc.prim_parcela, orc.seg_parcela, orc.terc_parcela, orc.total_dpvat, orc.desconto, orc.placas, orc.lacre, orc.outros, orc.srpr, orc.lacracao,
orc.placas_esp, orc.licenciamento, orc.multas, orc.vlr_bo, orc.total, orc.txvistoriamovel, (cp.ipva)ipva_cp, (cp.licenciamento)lic_cp, (cp.multas)multas_cp, codigo_cp, (orc.extras)Reembolso, m.observacao, m.solicitante
from os o
left join cliente c on c.codigo_cliente = o.cod_cliente
left join orcamento orc on orc.os = numero_os
left join cp_os on orc.os = num_os
left join contas_pagar cp on cod_cp = codigo_cp
left join manutencao m on m.os = numero_os
left join pendencias_os pdo on m.os = pdo.os
left join pendencias pd on pdo.codigo = pd.codigo
left join servicos s on o.cod_servico = s.codigo_servico

where pdo.codigo=120
and pdo.data_baixa='30-12-1899' """

    # Executa a consulta de CPs para aprovar
    cursor.execute(query_pedidoComp)
    # Resultado da consulta
    results = cursor.fetchall()

# Em caso de erro obtem e exibe o erro
except Exception as e:
    print(f"Erro na conexão: {e}")

# Cria e renomeia a planilha
nome_planilha = f'Pedido de compras {data_atual}.xlsx'
# Define o caminho completo da planilha
caminho_completo = os.path.join(caminho_planilha, nome_planilha)
# Define o cabeçalho da planilha
dados = {
         'Nome cliente': [],
         'Serviço': [],
         'OS': [],
         'Placa': [],
         'Honorarios': [],
         'Taxa Detran': [],
         'Validação': [],
         'Valor correio': [],
         'IPVA total': [],
         'Primeira parcela': [],
         'Segunda parcela': [],
         'Terceira parcela': [],
         'Total Dpvat': [],
         'Desconto': [],
         'Placas': [],
         'Lacre': [],
         'Outros': [],
         'Srpr': [],
         'Lacração': [],
         'Placas especial': [],
         'Licenciamento': [],
         'Multas': [],
         'Valor BO': [],
         'Total': [],
         'Taxa vostoria movel': [],
         'IPVA CP': [],
         'Licenciamento CP': [],
         'Multas CP': [],
         'Código CP': [],
         'Reembolso': [],
#         'Descrição': [],
         'Obrservação': [],
         'Solicitante': []
        }

# Crie um DataFrame com os dados
df = pd.DataFrame(dados)
# Salva o DataFrame na planilha no caminho especificado
df.to_excel(caminho_completo, index=False)

wb = load_workbook(caminho_completo)
ws = wb.active
wb.save(caminho_completo)

# Atribuir os valores ao cabeçalho
for col, valor in enumerate(dados, start=1):
    ws.cell(row=1, column=col, value=valor)

# Adicionar cor de fundo e estilo ao cabeçalho (linha 1)
cor_fundo = PatternFill(start_color="C3ECBC", end_color="C3ECBC", fill_type="solid")  # Amarelo
font_estilo = Font(bold=True, color="000000")  # Texto em negrito e preto

# Aplicar as mudanças nas células do cabeçalho (linha 1)
for col in range(1, len(dados) + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = cor_fundo  # Cor de fundo
    celula.font = font_estilo  # Estilo da fonte (negrito e preto)
# Salva a planilha após as mudanças    
wb.save(caminho_completo)
# Encontrar a primeira linha vazia
linha_vazia = df[df.isnull().all(axis=1)].index
if not linha_vazia.empty:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = linha_vazia[0] + 2
else:
    # Se não achar linha vazia pula o cabeçalho
    primeira_linha_vazia = len(df) + 2
    wb = load_workbook(caminho_completo)
    ws = wb.active
    # Percorres todas as linhas e acha a linha vazia
    for i, row in enumerate(results, start=primeira_linha_vazia):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
wb.save(caminho_completo)

# Iterar sobre as colunas da planilha (começando da última coluna para evitar problemas ao excluir)
for col in range(ws.max_column, 0, -1):  # Começa da última coluna e vai até a primeira
    # Verificar se todos os valores da coluna (exceto o cabeçalho) são 0 ou vazios
    all_zero_or_empty = True
    for row in range(2, ws.max_row + 1):  # Começa da linha 2 para ignorar o cabeçalho
        cell_value = ws.cell(row=row, column=col).value
        if cell_value not in (0, None):  # Verifica se o valor não é 0 nem vazio (None)
            all_zero_or_empty = False
            break
    # Exclui a coluna se todos os valores forem 0 ou vazios
    if all_zero_or_empty:
        ws.delete_cols(col)
# Salva as alterações feitas na planilha     
wb.save(caminho_completo)


""""